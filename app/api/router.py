"""HTTP 路由 —— 8 大功能 + 用户物料库/公共库/收藏/发布 + 多模态内容审核。只依赖 service(+组合根 deps)。"""
from __future__ import annotations
import uuid
import time
import hashlib
import threading
import zipfile
import datetime
import os
import re
import tempfile
import io
import json
import asyncio
import ipaddress
import socket
from pathlib import Path
from urllib.parse import urljoin, urlparse
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Header, Depends, Query
from fastapi.responses import StreamingResponse
from fastapi.concurrency import run_in_threadpool
from app.api import deps, schemas
from app.config import settings
from app.infrastructure.snowflake import next_id_str   # 规则主键:雪花 BIGINT 的字符串形态(PG 规范)
from app.domain.models import (MaterialType, AuditStatus, Material, AuditRule, User,
                               AuditTask, JobStatus, Project, TextSourceType,
                               MaterialSubmission, ShortDramaTask, VideoEditingTemplate, Requirement)
from app.domain.mp4 import parse_mp4_duration_ms
from app.service.material import MaterialNotFound
from app.service.user import InvalidCredentials, DuplicateName
from app.service.authorization import PermissionDenied
from app.service.video_compat import transcode_hevc_to_h264

router = APIRouter()
_UPLOAD_SCOPE_RE = re.compile(r"[^a-z0-9/_-]+")
_UPLOAD_ID_RE = re.compile(r"^[a-f0-9]{32}$")
_UPLOAD_PROGRESS: dict[str, dict] = {}
_UPLOAD_PROGRESS_LOCK = threading.Lock()


def _set_upload_progress(upload_id: str, user_id: str, **values) -> None:
    if not upload_id:
        return
    now = time.time()
    with _UPLOAD_PROGRESS_LOCK:
        # 进度只需短期保留；顺手清理过期任务，避免常驻进程无限增长。
        expired = [key for key, item in _UPLOAD_PROGRESS.items()
                   if now - float(item.get("updated_at", now)) > 3600]
        for key in expired:
            _UPLOAD_PROGRESS.pop(key, None)
        current = _UPLOAD_PROGRESS.get(upload_id, {"upload_id": upload_id, "user_id": user_id})
        current.update(values)
        current["updated_at"] = now
        _UPLOAD_PROGRESS[upload_id] = current


def _upload_file_size(fileobj) -> int:
    try:
        pos = fileobj.tell()
        fileobj.seek(0, 2)
        size = int(fileobj.tell())
        fileobj.seek(pos)
        return max(0, size)
    except Exception:
        return 0


def _user(authorization: str | None = Header(default=None)):
    return deps.current_user(authorization)


def _require_auth(user: dict) -> None:
    """必须是已登录用户(非游客)。"""
    if user["role"] == "guest":
        raise HTTPException(401, "请先登录")


def _require_perm(user: dict, permission: str) -> None:
    """RBAC 鉴权:按角色权限判定(后台 grant 即时生效)。无权限→403+审计。"""
    _require_auth(user)
    u = User(id=user["id"], name=user.get("name", ""), pwd_hash="", role=user["role"])
    try:
        deps.get_authz_service().authorize(u, permission)
    except PermissionDenied:
        raise HTTPException(403, "无权限执行该操作")


def _safe_upload_file_name(name: str) -> str:
    raw = (name or "").strip().replace("\\", "/")
    safe = os.path.basename(raw) or "upload.bin"
    return safe


def _safe_upload_scope(scope: str) -> str:
    s = (scope or "").strip().lower().replace("\\", "/")
    s = _UPLOAD_SCOPE_RE.sub("", s).strip("/")
    return s or "uploads"


def _can_view(user: dict, m) -> bool:
    """可查看/取签名URL:管理员 / 物主 / 已发布且审核通过(公共)。"""
    return (
        user["role"] == "admin"
        or m.owner_id == user["id"]
        or (m.is_public and m.audit_status == AuditStatus.PASS)
    )


def _preview_url(m) -> str:
    """卡片预览:图片→签名图;视频→OSS 截帧封面;声音/文字→无(前端显字形)。"""
    if not m.oss_key:
        return ""
    try:
        if m.type in (MaterialType.IMAGE, MaterialType.MEME, MaterialType.STYLE):
            return deps.storage.signed_url(m.oss_key)
        if m.type == MaterialType.VIDEO:
            return deps.storage.snapshot_url(m.oss_key)
    except Exception:
        return ""
    return ""


def _media_url(m) -> str:
    """审核卡片内联播放:图片/视频/声音都给真实文件签名 URL(前端直接 <img>/<video>/<audio>)。"""
    if not m.oss_key:
        return ""
    try:
        return deps.storage.signed_url(m.oss_key)
    except Exception:
        return ""


def _page_args(page: int, size: int) -> tuple[int, int]:
    """1 基 page/size → repo 的 offset/limit。"""
    return (page - 1) * size, size


def _page_out(items: list, total: int, page: int, size: int, key: str = "items") -> dict:
    """统一分页响应。count = 当页长度(向后兼容);翻页控件只认 total。"""
    return {"total": total, "page": page, "size": size, "count": len(items), key: items}


def _check_type(type: str | None) -> str | None:
    """校验物料类型(非法值 400,别静默返回空页)。"""
    if type:
        try:
            MaterialType(type)
        except ValueError:
            raise HTTPException(400, f"不支持的物料类型: {type}")
    return type or None


def _check_status(status: str | None) -> str | None:
    if status:
        try:
            AuditStatus(status)
        except ValueError:
            raise HTTPException(400, f"非法审核状态: {status}(应为 pass/review/block)")
    return status or None


def _owner_name(owner_id: str) -> str:
    """把 owner_id 解析成用户名(管理视图展示用);用户已删除 → 空(前端显示"已删除用户")。"""
    if not owner_id:
        return ""
    u = deps.user_repo.get(owner_id)
    return u.name if u else ""


def _check_can_upload_status(status: int | None) -> int | None:
    if status is None:
        return None
    if status not in (1, 2, 3):
        raise HTTPException(400, "非法可上传状态: 应为 1(可上传)、2(不可上传) 或 3(已修改)")
    return status


def _check_publish_status(status: int | None) -> int | None:
    if status is None:
        return None
    if status not in (1, 2):
        raise HTTPException(400, "非法发布状态: 应为 1(成功) / 2(失败)")
    return status


def _status_filter_arg(raw: str, *, kind: str) -> tuple[int | None, bool]:
    val = (raw or "").strip()
    if not val:
        return None, False
    if val == "__empty__":
        return None, True
    try:
        num = int(val)
    except ValueError:
        raise HTTPException(400, f"非法{kind}筛选值")
    checked = _check_can_upload_status(num) if kind == "可上传状态" else _check_publish_status(num)
    return checked, False


def _submission_permission(user: dict | None, submission: MaterialSubmission) -> str:
    if not user:
        return ""
    if user.get("role") == "admin":
        return "read_edit"
    joined_permission = getattr(submission, "_permission_type_hint", None)
    if joined_permission is not None:
        return joined_permission
    return deps.material_submission_repo.permission_of(submission.id, user.get("id", ""))


def _require_submission_access(user: dict, submission_id: str, *, edit: bool = False) -> MaterialSubmission:
    _require_auth(user)
    submission = deps.material_submission_repo.get(submission_id)
    if submission is None:
        raise HTTPException(404, "素材提报不存在")
    if user.get("role") == "admin":
        return submission
    permission_type = deps.material_submission_repo.permission_of(submission_id, user.get("id", ""))
    allowed = permission_type == "read_edit" if edit else permission_type in ("read", "read_edit")
    if not allowed:
        raise HTTPException(403, "无权访问该素材提报")
    return submission


def _require_deleted_submission_access(user: dict, submission_id: str,
                                       *, edit: bool = False) -> MaterialSubmission:
    _require_auth(user)
    submission = deps.material_submission_repo.get_deleted(submission_id)
    if submission is None:
        raise HTTPException(404, "回收站中的素材提报不存在或已完成 OSS 清理")
    if user.get("role") == "admin":
        return submission
    permission_type = deps.material_submission_repo.permission_of(submission_id, user.get("id", ""))
    allowed = permission_type == "read_edit" if edit else permission_type in ("read", "read_edit")
    if not allowed:
        raise HTTPException(403, "无权访问该素材提报")
    return submission


def _submission_out(s: MaterialSubmission, user: dict | None = None) -> dict:
    permission_type = _submission_permission(user, s)
    created_by_name = getattr(s, "_created_by_name_hint", None)
    if created_by_name is None:
        created_by_name = _owner_name(s.created_by)
    updated_by_name = getattr(s, "_updated_by_name_hint", None)
    if updated_by_name is None:
        updated_by_name = _owner_name(s.updated_by)
    video_url = ""
    decoded_video_url = ""
    if s.oss_key:
        try:
            video_url = deps.storage.signed_url(s.oss_key)
        except Exception:
            pass
    if s.decoded_oss_key:
        try:
            decoded_video_url = deps.storage.signed_url(s.decoded_oss_key)
        except Exception:
            pass
    return {
        "id": s.id,
        "team_name": s.team_name,
        "delivery_time": s.delivery_time,
        "drama_name": s.drama_name,
        "oss_key": s.oss_key,
        "video_url": video_url,
        "decoded_oss_key": s.decoded_oss_key,
        "decoded_video_url": decoded_video_url,
        "requires_decode": int(s.requires_decode),
        "video_file_name": s.video_file_name,
        "title_name": s.title_name,
        "episode_range": s.episode_range,
        "revision_comment": s.revision_comment,
        "can_upload_status": s.can_upload_status,
        "designated_upload_account_name": s.designated_upload_account_name,
        "upload_account_name": s.upload_account_name,
        "upload_date": s.upload_date,
        "publish_status": s.publish_status,
        "platform_reject_reason": s.platform_reject_reason,
        "platform_reject_attachments": list(s.platform_reject_attachments or []),
        "created_by": s.created_by,
        "created_by_name": created_by_name,
        "created_time": s.created_time,
        "updated_by": s.updated_by,
        "updated_by_name": updated_by_name,
        "updated_time": s.updated_time,
        "owner_name": created_by_name,
        "is_owner": bool(user and s.created_by == user.get("id")),
        "permission_type": permission_type,
        "can_read": permission_type in ("read", "read_edit"),
        "can_edit": permission_type == "read_edit",
    }


def _submission_decode_coordinator():
    coordinator = deps.submission_decode_coordinator
    if coordinator is None:
        raise HTTPException(503, "解码服务未配置 Redis")
    return coordinator


def _set_submission_decode_status(submission_id: str, status: str, stage: str, **values) -> None:
    coordinator = deps.submission_decode_coordinator
    if coordinator is None:
        return
    coordinator.set_status(
        submission_id,
        {
            "submission_id": submission_id,
            "status": status,
            "stage": stage,
            "updated_at": time.time(),
            **values,
        },
        ttl_seconds=settings.submission_decode_status_ttl_seconds,
    )


def _run_submission_decode_job(submission_id: str, source_key: str, file_name: str,
                               user_id: str, distributed_lock) -> None:
    import logging

    source_path = ""
    decoded_file = None
    decoded_key = ""
    committed = False
    try:
        _set_submission_decode_status(submission_id, "processing", "downloading", progress=3)
        source_fd, source_path = tempfile.mkstemp(suffix=os.path.splitext(file_name)[1] or ".mp4")
        os.close(source_fd)
        deps.storage.download_to_file(source_key, source_path)

        _set_submission_decode_status(submission_id, "processing", "transcoding", progress=10)
        def report_transcode_progress(percent: int) -> None:
            overall = min(90, 10 + round(max(0, min(100, percent)) * 0.8))
            _set_submission_decode_status(
                submission_id, "processing", "transcoding", progress=overall,
            )
        with open(source_path, "rb") as source:
            decoded_file = transcode_hevc_to_h264(
                source, progress_callback=report_transcode_progress,
            )
        if decoded_file is None:
            _set_submission_decode_status(
                submission_id, "not_required", "done", progress=100,
                source_key=source_key, decoded_oss_key="", error="",
            )
            return

        safe_name = _safe_upload_file_name(file_name or os.path.basename(source_key))
        stem = os.path.splitext(safe_name)[0] or "video"
        decoded_key = f"submissions/decoded/{uuid.uuid4().hex}-{stem}-h264.mp4"
        _set_submission_decode_status(submission_id, "processing", "uploading", progress=94)
        deps.storage.put_fileobj(decoded_key, decoded_file)

        if not deps.material_submission_repo.set_decoded_oss_key_if_current(
            submission_id, source_key, decoded_key, by=user_id,
        ):
            current = deps.material_submission_repo.get(submission_id)
            if current is not None and current.decoded_oss_key:
                deps.storage.delete(decoded_key)
                decoded_key = current.decoded_oss_key
                committed = True
            else:
                raise RuntimeError("源文件已被替换或提报已删除，兼容版未写入")
        else:
            committed = True
            try:
                deps.material_submission_repo.record_operation(
                    submission_id,
                    "update",
                    user_id,
                    [{"field": "decoded_oss_key", "before": "", "after": decoded_key}],
                )
            except Exception:
                logging.getLogger(__name__).exception("记录素材解码操作失败: submission_id=%s", submission_id)

        _set_submission_decode_status(
            submission_id, "done", "done", decoded_oss_key=decoded_key, error="", progress=100,
        )
    except Exception as exc:
        if decoded_key and not committed:
            try:
                deps.storage.delete(decoded_key)
            except Exception:
                pass
        logging.getLogger(__name__).exception("素材兼容版后台生成失败: submission_id=%s", submission_id)
        try:
            _set_submission_decode_status(
                submission_id, "error", "error", error=str(exc) or "兼容版生成失败", progress=0,
            )
        except Exception:
            logging.getLogger(__name__).exception("写入素材解码失败状态失败: submission_id=%s", submission_id)
    finally:
        if decoded_file is not None:
            decoded_file.close()
        if source_path:
            try:
                os.unlink(source_path)
            except FileNotFoundError:
                pass
        try:
            distributed_lock.release()
        except Exception:
            logging.getLogger(__name__).exception("释放素材解码 Redis 锁失败: submission_id=%s", submission_id)


def _submission_decode_status_out(submission: MaterialSubmission, user: dict) -> dict:
    if submission.decoded_oss_key:
        output = _submission_out(submission, user)
        return {
            "submission_id": submission.id,
            "status": "done",
            "stage": "done",
            "decoded_oss_key": submission.decoded_oss_key,
            "decoded_video_url": output.get("decoded_video_url", ""),
            "error": "",
            "progress": 100,
        }
    coordinator = _submission_decode_coordinator()
    try:
        status = coordinator.get_status(submission.id)
    except Exception as exc:
        raise HTTPException(503, "Redis 解码状态服务不可用") from exc
    return status or {
        "submission_id": submission.id,
        "status": "idle",
        "stage": "idle",
        "decoded_oss_key": "",
        "decoded_video_url": "",
        "error": "",
        "progress": 0,
    }


_SUBMISSION_OPERATION_FIELDS = (
    "team_name", "delivery_time", "drama_name", "oss_key", "decoded_oss_key", "requires_decode", "video_file_name",
    "title_name", "episode_range", "revision_comment", "can_upload_status",
    "designated_upload_account_name", "upload_account_name", "upload_date",
    "publish_status", "platform_reject_reason", "platform_reject_attachments",
)


def _submission_changes(before: MaterialSubmission, after: MaterialSubmission) -> list[dict]:
    changes = []
    for field in _SUBMISSION_OPERATION_FIELDS:
        old_value = getattr(before, field, None)
        new_value = getattr(after, field, None)
        if old_value != new_value:
            changes.append({"field": field, "before": old_value, "after": new_value})
    return changes


def _record_submission_operation(submission_id: str, action: str, user: dict,
                                 changes: list[dict]) -> None:
    deps.material_submission_repo.record_operation(submission_id, action, user.get("id", ""), changes)


def _submission_permission_changes(before: dict[str, str], after: dict[str, str]) -> list[dict]:
    changes = []
    for user_id in sorted(set(before) | set(after)):
        old_value = before.get(user_id, "")
        new_value = after.get(user_id, "")
        if old_value == new_value:
            continue
        changes.append({
            "field": "permission_type",
            "user_id": user_id,
            "user_name": _owner_name(user_id) or user_id,
            "before": old_value,
            "after": new_value,
        })
    return changes


def _submission_operation_asset_keys(submission_id: str) -> set[str]:
    keys: set[str] = set()
    for operation in deps.material_submission_repo.list_operations(submission_id):
        for change in operation.get("changes") or []:
            field = change.get("field")
            if field in ("oss_key", "decoded_oss_key"):
                values = (change.get("before"), change.get("after"))
            elif field == "platform_reject_attachments":
                values = [
                    key for side in (change.get("before"), change.get("after"))
                    for key in (side if isinstance(side, list) else [])
                ]
            else:
                continue
            keys.update(str(value).strip() for value in values if str(value or "").strip())
    return keys


def _template_status(value: str) -> str:
    status = (value or "").strip().lower()
    if status not in ("active", "inactive"):
        raise HTTPException(400, "Template status must be active or inactive")
    return status


def _template_oss_key(value: str, field_name: str) -> str:
    key = (value or "").strip()
    if not key:
        return ""
    if "://" in key or key.startswith("/") or key.startswith("\\"):
        raise HTTPException(400, f"{field_name} must be an OSS object key, not a URL or local path")
    return key


def _require_template_access(user: dict, template_id: str, *, edit: bool = False) -> VideoEditingTemplate:
    _require_auth(user)
    template = deps.video_editing_template_repo.get(template_id)
    if template is None:
        raise HTTPException(404, "Video-editing template not found")
    if edit and user.get("role") != "admin" and template.created_by != user.get("id"):
        raise HTTPException(403, "No permission to edit this video-editing template")
    return template


def _template_out(template: VideoEditingTemplate, user: dict | None = None) -> dict:
    can_edit = bool(user and (user.get("role") == "admin" or template.created_by == user.get("id")))
    return {
        "id": template.id,
        "name": template.name,
        "description": template.description,
        "reference_oss_key": template.reference_oss_key,
        "narration_voice": dict(template.narration_voice or {}),
        "bgm_oss_key": template.bgm_oss_key,
        "config": dict(template.config or {}),
        "status": template.status,
        "version": template.version,
        "created_by": template.created_by,
        "created_by_name": _owner_name(template.created_by),
        "created_time": template.created_time,
        "updated_by": template.updated_by,
        "updated_by_name": _owner_name(template.updated_by),
        "updated_time": template.updated_time,
        "can_edit": can_edit,
    }


def _mat_out(m, fav_ids: set | None = None, uid: str | None = None):
    return {
        "id": m.id, "type": m.type, "audit_status": m.audit_status,
        "oss_key": m.oss_key, "thumb": m.thumb, "description": m.description,
        "source_timecode": m.source_timecode, "owner_id": m.owner_id,
        "owner_name": _owner_name(m.owner_id),
        "is_public": m.is_public, "preview_url": _preview_url(m),
        "is_favorited": bool(fav_ids and m.id in fav_ids),
        "is_mine": bool(uid and m.owner_id == uid),
        "project_id": getattr(m, "project_id", ""),
        "tags": list(getattr(m, "tags", []) or []),
        "ai_summary": getattr(m, "ai_summary", ""),
        "ai_scenarios": list(getattr(m, "ai_scenarios", []) or []),
        "ai_emotions": list(getattr(m, "ai_emotions", []) or []),
        "ai_atmosphere": getattr(m, "ai_atmosphere", ""),
    }


# ── 物料管理(F1)──
@router.post("/materials")
def create_material(body: schemas.MaterialCreate, user: dict = Depends(_user)):
    _require_auth(user)
    m = deps.get_material_service().create(body.type, body.oss_key, b"", user["id"])
    deps.get_index_service().index_material(m)
    return _mat_out(m, uid=user["id"])


@router.post("/materials/upload")
async def upload_material(file: UploadFile = File(...), type: str = Form("image"), user: dict = Depends(_user)):
    """真文件上传:存 OSS + 落库(归属当前用户,状态 待审核)。"""
    _require_auth(user)
    try:
        mtype = MaterialType(type)
    except ValueError:
        raise HTTPException(400, f"不支持的物料类型: {type}")
    data = await file.read()
    key = f"materials/{uuid.uuid4().hex}-{file.filename}"
    m = deps.get_material_service().create(mtype, key, data, user["id"])
    deps.get_index_service().index_material(m)
    return _mat_out(m, uid=user["id"])


@router.get("/materials")
def list_materials(page: int = Query(1, ge=1), size: int = Query(24, ge=1, le=100),
                   type: str | None = None, status: str | None = None, q: str | None = None,
                   user: dict = Depends(_user)):
    """列出全部物料(含 review/block/他人)—— 仅管理员(审核队列用)。服务端分页/筛选。"""
    _require_perm(user, "materials.audit")
    off, lim = _page_args(page, size)
    items, total = deps.get_library_service().all(
        status=_check_status(status), type=_check_type(type), keyword=q or None,
        offset=off, limit=lim)
    return _page_out([_mat_out(m, uid=user["id"]) for m in items], total, page, size)


@router.get("/materials/{mid}")
def get_material(mid: str, user: dict = Depends(_user)):
    """取物料签名 URL。仅 管理员/物主/已发布过审(公共)可取,block/review/他人私有拒绝。"""
    m = deps.material_repo.get(mid)
    if m is None:
        raise HTTPException(404, "material not found")
    if not _can_view(user, m):
        raise HTTPException(403, "无权访问该物料")
    return {"id": mid, "signed_url": _media_url(m), "type": m.type, "preview_url": _preview_url(m)}


@router.get("/materials/{mid}/download")
def download_material(mid: str, user: dict = Depends(_user)):
    """下载物料文件 —— 仅"我的物料库"(我上传的 或 我收藏的)可下载;
    公共库里未收藏的物料不提供下载(先收藏进自己的库再下)。"""
    _require_auth(user)
    m = deps.material_repo.get(mid)
    if m is None:
        raise HTTPException(404, "material not found")
    in_my_library = m.owner_id == user["id"] or deps.favorites.has(user["id"], mid)
    if not (user["role"] == "admin" or in_my_library):
        raise HTTPException(403, "只能下载你物料库中的物料(公共物料请先收藏)")
    return {"download_url": deps.storage.download_url(m.oss_key)}


@router.post("/materials/{mid}/summarize")
def summarize_material(mid: str, user: dict = Depends(_user)):
    """按需生成 AI 摘要(重新解析物料 → 情绪/氛围/场景/标签)。仅物主或管理员。"""
    _require_auth(user)
    m = deps.material_repo.get(mid)
    if m is None:
        raise HTTPException(404, "material not found")
    if not (user["role"] == "admin" or m.owner_id == user["id"]):
        raise HTTPException(403, "只能给自己的物料生成摘要")
    deps.get_audit_service().summarize_material(m)
    return _mat_out(m, uid=user["id"])


@router.put("/materials/{mid}/tags")
def set_material_tags(mid: str, body: schemas.TagsIn, user: dict = Depends(_user)):
    """设置物料标签(项目分类)。仅物主或管理员。"""
    _require_auth(user)
    m = deps.material_repo.get(mid)
    if m is None:
        raise HTTPException(404, "material not found")
    if not (user["role"] == "admin" or m.owner_id == user["id"]):
        raise HTTPException(403, "只能修改自己物料的标签")
    m.tags = list(dict.fromkeys([t.strip() for t in body.tags if t.strip()]))[:12]
    deps.material_repo.save(m)
    return _mat_out(m, uid=user["id"])


@router.post("/materials/{mid}/set-audit")
def set_audit(mid: str, body: schemas.AuditSet, user: dict = Depends(_user)):
    """人工审核复核 —— 仅管理员(普通用户上传后等审核)。"""
    _require_perm(user, "materials.audit")
    try:
        new_status = AuditStatus(body.status)
    except ValueError:
        raise HTTPException(400, f"非法审核状态: {body.status}(应为 pass/review/block)")
    if new_status == AuditStatus.PROCESSING:               # 「审核中」是机器内部态,人工不可设
        raise HTTPException(400, "人工只能设 pass/review/block")
    m = deps.material_repo.get(mid)
    if m is None:
        raise HTTPException(404, "material not found")
    m.audit_status = new_status
    if new_status == AuditStatus.BLOCK:                     # 人工退回 → 记入退回历史(作品审核记录用)
        _record_reject(m, body.reason or "人工退回", "人工")
    deps.material_repo.save(m)
    return _mat_out(m, uid=user["id"])


def _record_reject(m, reason: str, by: str) -> None:
    """作品/物料被判 block 时追加一条退回记录。就地改 m.reject_events(调用方负责 save)。"""
    if not hasattr(m, "reject_events") or m.reject_events is None:
        m.reject_events = []
    m.reject_events.append({"ms": int(time.time() * 1000), "reason": (reason or "")[:200], "by": by})


@router.delete("/materials/{mid}")
def delete_material(mid: str, user: dict = Depends(_user)):
    """删除物料 —— 仅物主或管理员。"""
    _require_auth(user)
    m = deps.material_repo.get(mid)
    if m is None:
        raise HTTPException(404, "material not found")
    if not (user["role"] == "admin" or m.owner_id == user["id"]):
        raise HTTPException(403, "只能删除自己的物料")
    deps.get_material_service().delete(mid)
    return {"deleted": mid}


# ── 视频反解(F2/F5)──
@router.post("/videos")
def upload_video(body: schemas.VideoUpload, user: dict = Depends(_user)):
    _require_auth(user)
    vsvc = deps.get_video_service()
    job = vsvc.accept_upload(body.oss_key, body.size_bytes)
    deps.jobs[job.id] = {"status": "running", "materials": []}
    materials = vsvc.run_job(job, owner_id=user["id"])
    for m in materials:
        deps.get_index_service().index_material(m)
    deps.jobs[job.id] = {"status": job.status, "materials": [m.id for m in materials]}
    return {"job_id": job.id, "status": job.status, "material_count": len(materials)}


@router.post("/videos/upload")
async def upload_video_file(file: UploadFile = File(...), user: dict = Depends(_user)):
    """真视频上传:存 OSS → 受理 → 反解(归属当前用户)。反解在线程池执行,不阻塞事件循环。"""
    _require_auth(user)
    data = await file.read()
    key = f"videos/{uuid.uuid4().hex}-{file.filename}"
    await run_in_threadpool(deps.storage.put, key, data)
    vsvc = deps.get_video_service()
    job = vsvc.accept_upload(key, len(data))
    deps.jobs[job.id] = {"status": "running", "materials": []}
    materials = await run_in_threadpool(vsvc.run_job, job, user["id"])
    for m in materials:
        deps.get_index_service().index_material(m)
    deps.jobs[job.id] = {"status": job.status, "materials": [m.id for m in materials]}
    return {"job_id": job.id, "status": job.status,
            "materials": [_mat_out(m, uid=user["id"]) for m in materials]}


@router.get("/videos/{jid}")
def video_status(jid: str):
    job = deps.jobs.get(jid)
    if job is None:
        raise HTTPException(404, "job not found")
    return {"job_id": jid, **job}


# ── 多模态内容审核 ──
def _report_out(r) -> dict:
    trig = []
    for t in r.triggered:
        d = dict(t)
        fk = d.get("frame_oss_key")
        if fk:                                   # 命中的帧/图 → 给个签名 URL,报告里标红显示这张图
            try:
                d["frame_url"] = deps.storage.signed_url(fk)
            except Exception:
                d["frame_url"] = ""
        trig.append(d)
    # 把触发规则信息挂到对应 segment 上，前端可直接渲染"命中: 规则#5"
    trig_by_ms: dict[int, list[dict]] = {}
    for t in r.triggered:
        ms = t.get("begin_ms")
        if ms is not None:
            trig_by_ms.setdefault(ms, []).append(t)
    segs_out = []
    for s in r.segments:
        seg = {"source_type": s.source_type, "text": s.text, "begin_ms": s.begin_ms,
               "end_ms": s.end_ms, "frame_oss_key": s.frame_oss_key}
        # 匹配：triggered 的 begin_ms 落在 segment 时间范围内
        matched = []
        s_begin = s.begin_ms
        s_end = s.end_ms
        if s_begin is not None:
            for ms, items in trig_by_ms.items():
                if s_end is not None:
                    in_range = s_begin <= ms <= s_end
                else:
                    in_range = ms == s_begin
                if in_range:
                    for it in items:
                        matched.append({"rule_no": it.get("rule_no", 0),
                                        "rule_desc": (it.get("rule_desc") or "")[:80],
                                        "action": it.get("action", ""),
                                        "reason": (it.get("reason") or "")[:200]})
        seg["triggered_rules"] = matched
        segs_out.append(seg)
    return {
        "verdict": r.verdict, "summary": r.summary, "triggered": trig,
        "segments": segs_out,
    }


def _norm_level(v: str | None) -> str:
    """严格程度归一:literal=字面、regex=正则(不走大模型)保留原值;其余(含缺省/非法)→ metaphor(隐喻,安全默认)。"""
    return v if v in ("literal", "regex") else "metaphor"


def _norm_source_type(raw: str) -> str:
    """来源类型归一:逗号分隔多值,每部分必须是合法值;非法/空→ any"""
    parts = [p.strip() for p in (raw or "").split(",") if p.strip()]
    valid = [p for p in parts if p in _SOURCE_TYPES]
    return ",".join(valid) if valid else "any"


def _next_rule_no() -> int:
    """下一个规则编号:全局现有最大 no + 1(稳定、不复用、递增)。"""
    return max((getattr(r, "no", 0) for r in deps.rule_repo.list()), default=0) + 1


def _rule_out(r: AuditRule) -> dict:
    return {"id": r.id, "no": getattr(r, "no", 0), "source_type": r.source_type, "keywords": r.keywords,
            "condition": r.condition, "action": r.action, "enabled": r.enabled,
            "project_id": getattr(r, "project_id", ""),
            "guidance": getattr(r, "guidance", ""),
            "match_level": _norm_level(getattr(r, "match_level", "metaphor")),
            "regex": getattr(r, "regex", ""),
            "exceptions": getattr(r, "exceptions", [])}


def _project_out(p: Project) -> dict:
    return {"id": p.id, "name": p.name, "created_ms": p.created_ms}


def _task_out(t: AuditTask, in_training: bool = False) -> dict:
    # 任务裁定跟随物料现状:管理员在审核队列改判(pass/block)后,待审核页立即反映,
    # 不再停留在机审时的「待人工复核」。物料被删则退回任务存的裁定。
    verdict = t.verdict
    if t.material_id:
        m = deps.material_repo.get(t.material_id)
        if m is not None:
            verdict = getattr(m.audit_status, "value", m.audit_status)
    return {"id": t.id, "name": t.name, "material_type": t.material_type,
            "material_id": t.material_id, "status": t.status, "verdict": verdict,
            "report_id": t.report_id, "created_ms": t.created_ms,
            "report_generated_at": getattr(t, "report_generated_at", ""), "error": t.error,
            "video_kind": getattr(t, "video_kind", "material"),
            "project_id": getattr(t, "project_id", ""),
            "in_training": in_training}


def _new_task(owner_id: str, name: str, mtype: MaterialType, material_id: str, chash: str,
              video_kind: str = "material", project_id: str = "") -> AuditTask:
    task = AuditTask(id=uuid.uuid4().hex, owner_id=owner_id, name=name, material_type=mtype,
                     material_id=material_id, content_hash=chash, status=JobStatus.PENDING,
                     created_ms=int(time.time() * 1000), video_kind=video_kind, project_id=project_id)
    deps.task_repo.save(task)
    return task


def _fail_task(task: AuditTask, error: str) -> None:
    """审核失败:标失败 + 暴露原因 + 物料降级为 REVIEW(不删,方便重试)。
    不删 OSS/元数据 → 用户可从「待审核」页点「重试」重新跑审核;想彻底清除可手动删任务+物料。"""
    task.status = JobStatus.FAILED
    task.error = (error or "审核失败,请重试。")[:200]
    if task.material_id:
        try:
            m = deps.material_repo.get(task.material_id)
            if m is not None and m.audit_status == AuditStatus.PROCESSING:
                m.audit_status = AuditStatus.REVIEW
                deps.material_repo.save(m)
        except Exception:
            pass
    deps.task_repo.save(task)


def _finish_task(task: AuditTask, job, report, delete_on_fail: bool = True) -> None:
    # 机器只出 pass/review;退回历史只在人工拒绝(set-audit block)时记,机审不记。
    if job.status != JobStatus.DONE:                       # 审核没跑成(内部兜底转人工时把 job 标了 FAILED)
        if delete_on_fail:                                 # 首审失败 → 删没成功的物料 + 暴露原因(可重传)
            _fail_task(task, report.summary or "审核未完成,请重试。")
        else:                                              # 重判失败 → 只标失败,别删已入库的物料
            task.status = JobStatus.FAILED
            task.error = (report.summary or "重新审核失败,请重试。")[:200]
            deps.task_repo.save(task)
        return
    task.verdict = report.verdict.value
    task.status = JobStatus.DONE
    m = deps.material_repo.get(task.material_id)
    task.report_id = m.audit_report_id if m else ""
    task.report_generated_at = datetime.datetime.now(
        datetime.timezone(datetime.timedelta(hours=8))
    ).isoformat()
    deps.task_repo.save(task)


def _sync_task_after_recheck(mid: str, report) -> None:
    """按物料重判后,把关联的 AuditTask(若有)同步到新裁定/报告,避免「待审核任务」页与队列不一致。
    best-effort:没有任务就跳过;task_repo 无 material_id 索引,扫全量匹配(管理员单次动作,量小)。"""
    m = deps.material_repo.get(mid)
    new_rid = m.audit_report_id if m else ""
    generated_at = datetime.datetime.now(
        datetime.timezone(datetime.timedelta(hours=8))
    ).isoformat()
    for t in deps.task_repo.list_all():
        if t.material_id == mid:
            t.verdict = report.verdict.value
            t.status = JobStatus.DONE
            t.report_id = new_rid
            t.report_generated_at = generated_at
            t.error = ""
            deps.task_repo.save(t)


def _run_task_audit(task_id: str, text: str = "") -> None:
    """后台:对已建物料的任务跑审核,回写任务状态/裁定/报告(单条提交用)。"""
    task = deps.task_repo.get(task_id)
    if task is None:
        return
    task.status = JobStatus.RUNNING
    deps.task_repo.save(task)
    svc = deps.get_audit_service()
    m = deps.material_repo.get(task.material_id)
    job = svc.submit(task.material_type, oss_key=(m.oss_key if m else ""),
                     owner_id=task.owner_id, material_id=task.material_id,
                     video_kind=getattr(task, "video_kind", "material"),
                     project_id=getattr(task, "project_id", ""))
    try:
        report = svc.run(job, text)
        _finish_task(task, job, report)
    except Exception as e:
        _fail_task(task, str(e))   # 首审异常 → 删没成功的物料 + 暴露原因(可重传)


def _run_task_recheck(task_id: str) -> None:
    """后台:对已存报告用当前白名单/规则**只重判**(不重抽帧/转写),回写任务状态/裁定/报告。"""
    task = deps.task_repo.get(task_id)
    if task is None:
        return
    old = deps.report_repo.get(task.report_id) if task.report_id else None
    if old is None:
        return
    task.status = JobStatus.RUNNING
    task.error = ""
    deps.task_repo.save(task)
    svc = deps.get_audit_service()
    m = deps.material_repo.get(task.material_id)
    job = svc.submit(task.material_type, oss_key=(m.oss_key if m else ""),
                     owner_id=task.owner_id, material_id=task.material_id,
                     video_kind=getattr(task, "video_kind", "material"),
                     project_id=getattr(task, "project_id", ""))
    try:
        report = svc.recheck(job, old)
        _finish_task(task, job, report, delete_on_fail=False)   # 重判失败别删已入库的物料
    except Exception as e:
        task.status = JobStatus.FAILED
        task.error = str(e)[:200]
        deps.task_repo.save(task)


async def _batch_prepare_item(owner_id: str, name: str, data: bytes, mtype: MaterialType,
                               chash: str, video_kind: str, project_id: str,
                               fileobj=None) -> str | None:
    """批量内单条准备:上传 OSS + 建物料 + 时长检查 + 建任务,只提交「审核」到线程池(数据已释放)。
    在调用方循环内 await,完成后 data 引用即可释放,内存峰值仅当前单条。
    有 fileobj 时优先流式直传 OSS(免 data 二次拷贝);无则用 data 上传(zip 条目/语料)。
    返回 task_id;失败时 _fail_task 标失败并仍返回 task_id(前端可追踪失败原因)。"""
    svc = deps.get_audit_service()
    msvc = deps.get_material_service()
    task = _new_task(owner_id, name, mtype, "", chash,
                     video_kind=video_kind, project_id=project_id)
    try:
        if mtype == MaterialType.CORPUS:
            text = data.decode("utf-8", "ignore")
            m = Material(id=uuid.uuid4().hex, type=mtype, thumb="", source_timecode=0.0, embedding=[],
                         audit_status=AuditStatus.PROCESSING, source_job="", oss_key="",
                         description=text, owner_id=owner_id, content_hash=chash)
            deps.material_repo.save(m)
        else:
            text = ""
            key = f"materials/{uuid.uuid4().hex}-{name.rsplit('/', 1)[-1]}"
            if fileobj is not None:
                # 流式:从 file-like 对象分块直传 OSS,避免 data bytes 二次拷贝
                m = await run_in_threadpool(msvc.create_file, mtype, key, fileobj, owner_id, chash)
            else:
                m = await run_in_threadpool(msvc.create, mtype, key, data, owner_id, chash)
            # 物料视频 ≤20s 护栏(与单个上传一致):超时长→删物料+任务失败,不入库不审核;请改选「作品」
            if mtype == MaterialType.VIDEO and video_kind == "material":
                dur = parse_mp4_duration_ms(data)   # 优先内存解析,避免 OSS 回读
                if dur is None:
                    dur = await run_in_threadpool(deps.storage.video_duration_ms, m.oss_key)
                if dur is not None and dur > 20000:
                    await run_in_threadpool(msvc.delete, m.id)
                    task.status = JobStatus.FAILED
                    task.error = "物料视频需 ≤20 秒;请改选「作品」或裁剪后重传。"
                    deps.task_repo.save(task)
                    return task.id
            deps.get_index_service().index_material(m)
        if project_id and video_kind == "work" and mtype == MaterialType.VIDEO:   # 作品(视频)落项目
            m.project_id = project_id
            deps.material_repo.save(m)
        task.material_id = m.id
        deps.task_repo.save(task)
        # 只提交审核到有界池,不传 data —— 内存已释放
        deps.audit_pool.submit(_batch_run_audit, task.id, m.id, mtype, m.oss_key, text, owner_id)
        return task.id
    except Exception as e:
        _fail_task(task, str(e))
        return task.id


def _batch_run_audit(task_id: str, material_id: str, mtype: MaterialType, oss_key: str,
                     text: str, owner_id: str) -> None:
    """纯审核(线程池内跑):OSS 已上传,只跑 audit pipeline + 回写任务。"""
    svc = deps.get_audit_service()
    task = deps.task_repo.get(task_id)
    if task is None:
        return
    task.status = JobStatus.RUNNING
    deps.task_repo.save(task)
    kind = getattr(task, "video_kind", "material")
    pid = getattr(task, "project_id", "")
    try:
        job = svc.submit(mtype, oss_key=oss_key, owner_id=owner_id, material_id=material_id,
                         video_kind=kind, project_id=pid)
        report = svc.run(job, text)
        _finish_task(task, job, report)
    except Exception as e:
        _fail_task(task, str(e))


# 去重要原子:检查「库内已有」+「同内容正在处理中」并登记,防并发/连点重复提交(检查-建库非原子的竞态)
_dedup_lock = threading.Lock()
_inflight_hashes: set = set()


def _task_for_material(mid: str):
    """按 material_id 找它的审核任务(量小,遍历可接受)。"""
    if not mid:
        return None
    return next((t for t in deps.task_repo.list_all() if t.material_id == mid), None)


def _purge_if_dead(m) -> bool:
    """m 是否「死上传」(有对应任务且已 failed:审核没成功)。是→清掉物料+失败任务并返回 True。
    兜底覆盖历史残留 + 删物料与去重之间的竞态,别让没成功的上传永久挡住重传。"""
    t = _task_for_material(m.id)
    if t is None or t.status != JobStatus.FAILED:
        return False
    try:
        deps.get_material_service().delete(m.id)
    except Exception:
        pass
    try:
        deps.task_repo.delete(t.id)
    except Exception:
        pass
    return True


def _dedup_reserve(owner_id: str, chash: str):
    """原子登记。返回 (已存在物料 or None, 是否重复)。重复=库内已有 或 同内容正在处理中。
    库内命中若是「死上传」(审核没成功的残留)→ 清掉、当作不重复放行(不挡重传)。"""
    with _dedup_lock:
        existing = deps.material_repo.by_content_hash(owner_id, chash)
        if existing is not None and not _purge_if_dead(existing):
            return existing, True
        if (owner_id, chash) in _inflight_hashes:
            return None, True
        _inflight_hashes.add((owner_id, chash))
        return None, False


def _dedup_release(owner_id: str, chash: str) -> None:
    with _dedup_lock:
        _inflight_hashes.discard((owner_id, chash))


def _resolve_project(video_kind: str, project_id: str) -> str:
    """作品(video_kind=work)必须选一个存在的项目;非作品不带项目。归一化并校验;非法→400。"""
    project_id = (project_id or "").strip()
    if video_kind == "work":
        if not project_id:
            raise HTTPException(400, "作品必须选择所属项目。")
        if deps.project_repo.get(project_id) is None:
            raise HTTPException(400, "所选项目不存在,请刷新后重试。")
        return project_id
    return ""


@router.post("/audit/submit")
async def audit_submit(type: str = Form("image"), content: str = Form(""),
                       video_kind: str = Form("material"), project_id: str = Form(""),
                       file: UploadFile = File(None), user: dict = Depends(_user)):
    """审核入口:上传即成功、可立刻再提交;审核异步跑,统一到「待审核」页看状态。
    同一用户库内按内容 MD5 去重,重复不再上传。视频分 物料(material,≤20s,抽帧入库)/ 作品(work,仅扫描)。"""
    _require_auth(user)
    try:
        mtype = MaterialType(type)
    except ValueError:
        raise HTTPException(400, f"不支持的类型: {type}")
    owner = user["id"]
    video_kind = video_kind if video_kind in ("material", "work") else "material"
    project_id = _resolve_project(video_kind, project_id)   # 作品必须选存在的项目

    # 文字:内容 hash 原子去重(防连点/并发)→ 建语料物料 → 异步审核
    if mtype == MaterialType.CORPUS:
        text = content.strip()
        if not text:
            raise HTTPException(400, "文字内容不能为空")
        chash = hashlib.md5(text.encode("utf-8")).hexdigest()
        existing, is_dup = _dedup_reserve(owner, chash)
        if is_dup:
            return {"status": "duplicate", "material_id": existing.id if existing else "",
                    "message": "这段文字你已提交过,未重复。"}
        try:
            m = Material(id=uuid.uuid4().hex, type=mtype, thumb="", source_timecode=0.0, embedding=[],
                         audit_status=AuditStatus.PROCESSING, source_job="", oss_key="",
                         description=text, owner_id=owner, content_hash=chash)
            deps.material_repo.save(m)
            task = _new_task(owner, "文字审核", mtype, m.id, chash)
            deps.audit_pool.submit(_run_task_audit, task.id, text)   # 提交到有界审核池(超上限排队=背压)
            return {"status": "submitted", "task_id": task.id, "material_id": m.id}
        finally:
            _dedup_release(owner, chash)

    # 文件:校验(格式/大小)→ 读字节 → 原子去重 → 存 OSS 建物料(此步=上传成功)→ 异步审核
    if file is None:
        raise HTTPException(400, "缺少文件")
    fname = file.filename or "文件"
    terr = _type_error(fname, mtype.value)                 # 不正确的物料/作品 → 明确提示
    if terr:
        raise HTTPException(400, terr)
    if getattr(file, "size", None) and file.size > _MAX_UPLOAD:   # 按 Content-Length 粗检,省得白传 1GB 再拒
        raise HTTPException(413, _size_error(file.size))
    data = await file.read()
    if not data:
        raise HTTPException(400, "文件是空的,请重新选择。")
    if len(data) > _MAX_UPLOAD:                             # 按真实字节数定检
        raise HTTPException(413, _size_error(len(data)))
    chash = hashlib.md5(data).hexdigest()
    existing, is_dup = _dedup_reserve(owner, chash)
    if is_dup:
        return {"status": "duplicate", "material_id": existing.id if existing else "",
                "message": f"「{fname}」已在你的库中,未重复上传。"}
    try:
        key = f"audit/{uuid.uuid4().hex}-{fname}"
        # 流式上传:seek 回文件头,OSS 从 file-like 对象分块读取,避免全量 bytes 再次拷贝
        await file.seek(0)
        m = await run_in_threadpool(deps.get_material_service().create_file,
                                     mtype, key, file.file, owner, chash)
        if project_id and mtype == MaterialType.VIDEO:     # 作品(视频)落项目(队列按 Material.project_id 分栏/筛)
            m.project_id = project_id
            deps.material_repo.save(m)
        # 物料视频强制 ≤20 秒(作品不限);优先从内存 data 解析 MP4 时长,失败回退 OSS
        if mtype == MaterialType.VIDEO and video_kind == "material":
            dur = parse_mp4_duration_ms(data)
            if dur is None:
                dur = await run_in_threadpool(deps.storage.video_duration_ms, m.oss_key)
            if dur is not None and dur > 20000:
                await run_in_threadpool(deps.get_material_service().delete, m.id)  # 删 OSS + 元数据
                return {"status": "too_long",
                        "message": f"物料视频需 ≤20 秒,当前约 {round(dur/1000)} 秒;请改选「作品」或裁剪后再传。"}
        deps.get_index_service().index_material(m)
        task = _new_task(owner, fname, mtype, m.id, chash, video_kind=video_kind, project_id=project_id)
        deps.audit_pool.submit(_run_task_audit, task.id, "")   # 提交到有界审核池
        return {"status": "submitted", "task_id": task.id, "material_id": m.id}
    except HTTPException:
        raise
    except Exception as _e:                                 # OSS 上传/建库/索引失败 → 友好提示,不抛 500
        import traceback
        traceback.print_exc()
        raise HTTPException(502, f"上传到存储或建库失败,请稍后重试。({_e})")
    finally:
        _dedup_release(owner, chash)


# ── 批量上传(ZIP 解包 / 文件夹多文件)──
_EXT_TYPE = {
    "jpg": "image", "jpeg": "image", "png": "image", "gif": "image", "webp": "image", "bmp": "image",
    "mp4": "video", "mov": "video", "mkv": "video", "avi": "video", "webm": "video", "flv": "video",
    "mp3": "audio", "wav": "audio", "m4a": "audio", "aac": "audio", "flac": "audio", "ogg": "audio",
    "txt": "corpus", "md": "corpus",
}


def _infer_type(name: str):
    ext = name.rsplit(".", 1)[-1].lower() if "." in name else ""
    return _EXT_TYPE.get(ext)


_MAX_UPLOAD = 1024 * 1024 * 1024   # 单文件上限 1 GB
_TYPE_CN = {"image": "图片", "video": "视频", "audio": "声音", "corpus": "文字",
            "meme": "表情包", "style": "风格", "music": "音乐"}


def _size_error(n: int) -> str:
    return f"文件不能超过 1GB(当前约 {round(n / 1024 / 1024)} MB),请压缩或裁剪后再传。"


# 音频文件既可当「声音」也可当「音乐/歌曲」——同一媒体家族,语义标签由用户选(音乐才走联网搜档案)
_TYPE_FAMILY = {"audio": "audio", "music": "audio"}


def _type_error(filename: str, mtype_value: str) -> str | None:
    """上传的是否『正确的物料/作品』:格式支持 + 与所选类型相符。返回中文错误(None=OK)。
    声音/音乐同属音频家族:音频文件选「音乐」也放行(歌曲→联网搜情绪/场景)。"""
    inferred = _infer_type(filename)
    if inferred is None:
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        return f"不支持的文件格式{('「.' + ext + '」') if ext else ''};请上传 图片/视频/声音/音乐 文件(文本用「文字」粘贴)。"
    if _TYPE_FAMILY.get(inferred, inferred) != _TYPE_FAMILY.get(mtype_value, mtype_value):
        return (f"文件与所选类型不符:你选了「{_TYPE_CN.get(mtype_value, mtype_value)}」,"
                f"但「{filename}」是{_TYPE_CN.get(inferred, inferred)}文件;请改选类型或换文件。")
    return None


def _zip_entry_ok(info) -> bool:
    """zip 条目跳过目录、隐藏文件、__MACOSX。"""
    if info.is_dir():
        return False
    base = info.filename.rsplit("/", 1)[-1]
    return not (info.filename.startswith("__MACOSX") or base.startswith("."))


@router.post("/audit/batch")
async def audit_batch(files: list[UploadFile] = File(...),
                      video_kind: str = Form("material"), project_id: str = Form(""),
                      user: dict = Depends(_user)):
    """批量:多文件(文件夹拖拽)或单个 zip(自动解包)。逐个上传+审核,状态在「待审核」页看。
    视频统一按顶部 tab 的 video_kind(material/work)分类(不再按时长自动猜);物料视频仍需 ≤20s。
    同一用户库内按内容 MD5 去重(库内已有 + 批内重复都跳过);不支持的扩展名也跳过。
    zip 文件直接从磁盘流式解压,不把整个压缩包读入内存。"""
    _require_auth(user)
    owner = user["id"]
    video_kind = video_kind if video_kind in ("material", "work") else "material"
    project_id = _resolve_project(video_kind, project_id)   # 作品批量必须选存在的项目
    task_ids: list[str] = []
    skipped_big = skipped_type = skipped_dup = 0
    seen: set = set()
    count = 0

    for f in files:
        fname = f.filename or "file"

        if fname.lower().endswith(".zip"):
            # 流式解压:直接从 UploadFile 的 SpooledTemporaryFile 读,不把整个 zip 加载到内存
            await f.seek(0)
            try:
                with zipfile.ZipFile(f.file) as z:
                    for info in z.infolist():
                        if not _zip_entry_ok(info):
                            continue
                        if count >= 200:
                            break
                        name = info.filename
                        if info.file_size > _MAX_UPLOAD:
                            skipped_big += 1
                            continue
                        t = _infer_type(name)
                        if t is None:
                            skipped_type += 1
                            continue
                        # 逐条读取,内存只持有当前这条的 bytes
                        data = z.read(info)
                        if not data:
                            continue
                        chash = hashlib.md5(data).hexdigest()
                        if chash in seen or deps.material_repo.by_content_hash(owner, chash) is not None:
                            skipped_dup += 1
                            continue
                        seen.add(chash)
                        mtype = MaterialType(t)
                        tid = await _batch_prepare_item(
                            owner, name.rsplit("/", 1)[-1], data, mtype, chash,
                            video_kind, project_id)
                        if tid:
                            task_ids.append(tid)
                            count += 1
            except Exception:
                pass   # 损坏/非 zip 文件 → 跳过,不阻塞批量
        else:
            # 非 zip 单文件:读入哈希 + 去重,然后 seek 回文件头流式传 OSS(免 data 二次拷贝)
            data = await f.read()
            if not data or len(data) > _MAX_UPLOAD:
                skipped_big += 1 if data else 0
                continue
            t = _infer_type(fname)
            if t is None:
                skipped_type += 1
                continue
            chash = hashlib.md5(data).hexdigest()
            if chash in seen or deps.material_repo.by_content_hash(owner, chash) is not None:
                skipped_dup += 1
                continue
            seen.add(chash)
            mtype = MaterialType(t)
            await f.seek(0)   # seek 回文件头,OSS 从 file-like 对象分块直传,不传 data bytes
            tid = await _batch_prepare_item(
                owner, fname.rsplit("/", 1)[-1], data, mtype, chash,
                video_kind, project_id, fileobj=f.file)
            if tid:
                task_ids.append(tid)
                count += 1
            if count >= 200:
                break

    skipped = skipped_big + skipped_type + skipped_dup
    _sk = {"skipped": skipped, "skipped_big": skipped_big,
           "skipped_type": skipped_type, "skipped_dup": skipped_dup}
    if not task_ids:
        return {"status": "done", "created": 0, **_sk, "task_ids": []}
    return {"status": "submitted", "created": len(task_ids), **_sk,
            "task_ids": task_ids}


# ── 审核规则后台(管理员)——放在 /audit/{job_id} 之前,避免 rules 被当作 job_id ──
@router.get("/audit/rules")
def list_audit_rules(project: str | None = None, user: dict = Depends(_user)):
    """列规则。project 缺省=全部;project=""=只看标准/全局;project=P=只看该项目规则。"""
    _require_perm(user, "audit.rules")
    rules = deps.rule_repo.list()
    if project is not None:
        rules = [r for r in rules if getattr(r, "project_id", "") == project]
    return {"rules": [_rule_out(r) for r in rules]}


@router.post("/audit/rules")
def add_audit_rule(body: schemas.RuleIn, user: dict = Depends(_user)):
    _require_perm(user, "audit.rules")
    action = body.action if body.action in ("block", "review") else "block"
    project_id = (body.project_id or "").strip()
    if project_id and deps.project_repo.get(project_id) is None:
        raise HTTPException(400, "所选项目不存在。")
    st = _norm_source_type(body.source_type)
    rule = AuditRule(id=next_id_str(), no=_next_rule_no(), source_type=st,
                     keywords=[k for k in body.keywords if k.strip()], condition=body.condition.strip(),
                     action=action, enabled=True, created_by=user["id"], project_id=project_id,
                     guidance=(body.guidance or "").strip(), match_level=_norm_level(body.match_level),
                     regex=(body.regex or "").strip())
    deps.rule_repo.add(rule, by=user["id"])
    return _rule_out(rule)


@router.put("/audit/rules/{rule_id}")
def update_audit_rule(rule_id: str, body: schemas.RuleIn, user: dict = Depends(_user)):
    """编辑已有规则:按 id 覆盖(保留 id/created_by/enabled,其余按请求更新)。归一化/校验同新增。"""
    _require_perm(user, "audit.rules")
    existing = next((r for r in deps.rule_repo.list() if r.id == rule_id), None)
    if existing is None:
        raise HTTPException(404, "规则不存在。")
    action = body.action if body.action in ("block", "review") else "block"
    project_id = (body.project_id or "").strip()
    if project_id and deps.project_repo.get(project_id) is None:
        raise HTTPException(400, "所选项目不存在。")
    updated = AuditRule(id=rule_id, no=getattr(existing, "no", 0) or _next_rule_no(),
                        source_type=_norm_source_type(body.source_type),
                        keywords=[k for k in body.keywords if k.strip()], condition=body.condition.strip(),
                        action=action, enabled=existing.enabled, created_by=existing.created_by,
                        project_id=project_id, guidance=(body.guidance or "").strip(),
                        match_level=_norm_level(body.match_level), regex=(body.regex or "").strip(),
                        exceptions=getattr(existing, "exceptions", []))   # 编号/例外不随编辑丢失
    deps.rule_repo.add(updated, by=user["id"])   # 按 id 覆盖(upsert)
    return _rule_out(updated)


_SOURCE_TYPES = {"any"} | {t.value for t in TextSourceType}


@router.post("/audit/rules/parse")
def parse_audit_rules(body: schemas.RuleParseIn, user: dict = Depends(_user)):
    """粘贴整篇「卡审/审核标准」文案 → 大模型拆成结构化规则草案(预览用,不落库)。
    需先选定项目作用域(作品规则都归属某个项目)。前端预览可删个别条后再走 /audit/rules/bulk 落库。"""
    _require_perm(user, "audit.rules")
    project_id = (body.project_id or "").strip()
    if not project_id:
        raise HTTPException(400, "请先选择规则所属的项目。")
    if deps.project_repo.get(project_id) is None:
        raise HTTPException(400, "所选项目不存在,请刷新后重试。")
    text = (body.text or "").strip()
    if not text:
        raise HTTPException(400, "请粘贴要解析的审核文案。")
    drafts = deps.get_audit_service().parse_rules(text)
    if not drafts:
        raise HTTPException(422, "没能从这段文案里解析出规则,请检查内容后重试。")
    return {"rules": drafts, "project_id": project_id}


@router.post("/audit/rules/compile-regex")
def compile_rule_regex(body: schemas.RegexCompileIn, user: dict = Depends(_user)):
    """正则规则:把管理员的自然语言描述交大模型编译成 {keywords, regex}(预览用,不落库)。
    只在建/编辑规则时用一次;审核时用编译出的正则纯匹配、不再调大模型。管理员可在前端再手改 regex。"""
    _require_perm(user, "audit.rules")
    text = (body.text or "").strip()
    if not text:
        raise HTTPException(400, "请先填写要拦的内容(自然语言)。")
    out = deps.get_audit_service().compile_regex(text)
    if not out.get("regex") and not out.get("keywords"):
        raise HTTPException(422, "没能从这段描述编译出正则,请换个说法重试。")
    return out


@router.post("/audit/rules/bulk")
def bulk_add_audit_rules(body: schemas.RulesBulkIn, user: dict = Depends(_user)):
    """把预览确认后的规则草案批量落库到指定项目作用域;逐条归一化并跳过空规则。"""
    _require_perm(user, "audit.rules")
    project_id = (body.project_id or "").strip()
    if not project_id:
        raise HTTPException(400, "请先选择规则所属的项目。")
    if deps.project_repo.get(project_id) is None:
        raise HTTPException(400, "所选项目不存在,请刷新后重试。")
    created: list[AuditRule] = []
    for d in body.rules:
        st = _norm_source_type(d.source_type)
        kws = list(dict.fromkeys(k.strip() for k in d.keywords if k.strip()))
        cond = (d.condition or "").strip()
        if not kws and not cond:
            continue    # 空规则(无词无条件)—— 无法命中,跳过
        action = d.action if d.action in ("block", "review") else "review"
        rule = AuditRule(id=next_id_str(), no=_next_rule_no(), source_type=st, keywords=kws,
                         condition=cond, action=action, enabled=True,
                         created_by=user["id"], project_id=project_id,
                         match_level=_norm_level(getattr(d, "match_level", "metaphor")))
        deps.rule_repo.add(rule, by=user["id"])   # 立刻落库 → 下一条 _next_rule_no 见到它、编号递增
        created.append(rule)
    return {"created": len(created), "rules": [_rule_out(r) for r in created]}


@router.delete("/audit/rules/{rule_id}")
def delete_audit_rule(rule_id: str, user: dict = Depends(_user)):
    _require_perm(user, "audit.rules")
    deps.rule_repo.delete(rule_id, by=user["id"])
    return {"deleted": rule_id}


_SYNTHETIC_RULE_IDS = {"", "blockword", "content-safety"}


@router.post("/audit/rules/{rule_id}/exceptions")
def add_rule_exception(rule_id: str, body: schemas.RuleExceptionIn, user: dict = Depends(_user)):
    """审核员「忽略这条」→ 把这段命中内容记为该规则的可放行例外(喂回语义判定,后续同类放行)。
    仅对真规则生效(禁词/内容安全等合成命中不走这里:禁词去删词、内容安全用白名单)。"""
    _require_perm(user, "audit.rules")
    if rule_id in _SYNTHETIC_RULE_IDS:
        raise HTTPException(400, "该命中不是规则命中,无法记为规则例外(禁词请去禁词库删词,内容安全请用白名单)。")
    rule = next((r for r in deps.rule_repo.list() if r.id == rule_id), None)
    if rule is None:
        raise HTTPException(404, "规则不存在。")
    text = (body.text or "").strip() or (body.note or "").strip()   # 无定位文本时退回用 AI 原因
    if not text:
        raise HTTPException(400, "例外内容为空。")
    if not hasattr(rule, "exceptions") or rule.exceptions is None:
        rule.exceptions = []
    rule.exceptions.append({"text": text[:500], "note": (body.note or "").strip()[:500],
                            "by": user["id"], "ms": int(time.time() * 1000)})
    deps.rule_repo.add(rule, by=user["id"])
    return _rule_out(rule)


@router.delete("/audit/rules/{rule_id}/exceptions")
def delete_rule_exception(rule_id: str, index: int, user: dict = Depends(_user)):
    """删掉规则的第 index 条例外(撤销误标)。"""
    _require_perm(user, "audit.rules")
    rule = next((r for r in deps.rule_repo.list() if r.id == rule_id), None)
    if rule is None:
        raise HTTPException(404, "规则不存在。")
    exc = getattr(rule, "exceptions", None) or []
    if 0 <= index < len(exc):
        exc.pop(index)
        rule.exceptions = exc
        deps.rule_repo.add(rule, by=user["id"])
    return _rule_out(rule)


# ── 待审核任务(异步审核状态,统一呈现在「待审核」页;用户看自己的,管理员看全部;支持分页+按项目筛选)──
@router.get("/audit/tasks")
def list_audit_tasks(page: int = Query(1, ge=1), size: int = Query(20, ge=1, le=100),
                    project_id: str = Query(""), name: str = Query(""),
                    user: dict = Depends(_user)):
    _require_auth(user)
    off, lim = _page_args(page, size)
    is_admin = user["role"] == "admin"
    tasks = (deps.task_repo.list_all(project_id=project_id, name=name, offset=off, limit=lim) if is_admin
             else deps.task_repo.list_for(user["id"], project_id=project_id, name=name, offset=off, limit=lim))
    total = deps.task_repo.count_all(project_id=project_id, name=name) if is_admin \
        else deps.task_repo.count_for(user["id"], project_id=project_id, name=name)
    training_materials_by_project: dict[str, set[str]] = {}
    for t in tasks:
        pid = getattr(t, "project_id", "") or ""
        mid = getattr(t, "material_id", "") or ""
        if not pid or not mid:
            continue
        if pid in training_materials_by_project:
            continue
        examples = deps.get_training_service().list_examples(pid)
        training_materials_by_project[pid] = {e.material_id for e in examples}
    return _page_out([
        _task_out(
            t,
            in_training=((getattr(t, "material_id", "") or "") in
                         training_materials_by_project.get(getattr(t, "project_id", "") or "", set()))
        )
        for t in tasks
    ], total, page, size, key="tasks")


@router.get("/audit/tasks/{task_id}")
def get_audit_task(task_id: str, user: dict = Depends(_user)):
    _require_auth(user)
    t = deps.task_repo.get(task_id)
    if t is None:
        raise HTTPException(404, "task not found")
    if not (user["role"] == "admin" or t.owner_id == user["id"]):
        raise HTTPException(403, "无权查看该任务")
    report = deps.report_repo.get(t.report_id) if t.report_id else None
    return {**_task_out(t), "report": _report_out(report) if report else None}


@router.delete("/audit/tasks/{task_id}")
def delete_audit_task(task_id: str, user: dict = Depends(_user)):
    _require_auth(user)
    t = deps.task_repo.get(task_id)
    if t is not None and not (user["role"] == "admin" or t.owner_id == user["id"]):
        raise HTTPException(403, "无权删除该任务")
    deps.task_repo.delete(task_id)
    return {"deleted": task_id}


@router.post("/audit/tasks/{task_id}/recheck")
def recheck_audit_task(task_id: str, user: dict = Depends(_user)):
    """加白/改规则后,用当前白名单重新判定该任务(画面用当前 vision 提示词重新反解,不重抽帧/转写)。改判需审核权限。"""
    _require_perm(user, "materials.audit")
    t = deps.task_repo.get(task_id)
    if t is None:
        raise HTTPException(404, "task not found")
    if t.status != JobStatus.DONE or not t.report_id:
        raise HTTPException(400, "仅可对已完成且有报告的任务重新审核")
    t.status = JobStatus.RUNNING   # 同步置「审核中」→ 前端立刻看到并开始轮询(消除竞态)
    t.error = ""
    deps.task_repo.save(t)
    deps.audit_pool.submit(_run_task_recheck, t.id)   # 提交到有界审核池
    return {"status": "rechecking", "id": t.id}


@router.post("/audit/tasks/{task_id}/retry")
def retry_audit_task(task_id: str, user: dict = Depends(_user)):
    """对失败的任务重新跑**完整**审核(重新抽帧/转写/反解)。需审核权限。
    与 recheck 区别:recheck 复用已存报告只重判;retry 从零跑全流程。"""
    _require_perm(user, "materials.audit")
    t = deps.task_repo.get(task_id)
    if t is None:
        raise HTTPException(404, "task not found")
    if t.status != JobStatus.FAILED:
        raise HTTPException(400, "仅可重试失败的任务")
    m = deps.material_repo.get(t.material_id) if t.material_id else None
    if m is None:
        raise HTTPException(400, "物料已被删除,请重新上传")
    t.status = JobStatus.RUNNING   # 同步置「审核中」→ 前端立刻看到并开始轮询
    t.error = ""
    deps.task_repo.save(t)
    deps.audit_pool.submit(_run_task_audit, t.id, "")
    return {"status": "retrying", "id": t.id}


@router.post("/materials/{mid}/recheck")
def recheck_material(mid: str, user: dict = Depends(_user)):
    """审核队列里对单条物料「按最新规则重新审核」:画面用当前 vision 提示词重新反解,
    口播/原文复用已存 segments;只用**当前**规则重跑三波级联 → 回写报告 + 物料状态 + 关联任务,同步返回新报告(标红即刷新)。"""
    _require_perm(user, "materials.audit")
    m = deps.material_repo.get(mid)
    if m is None:
        raise HTTPException(404, "material not found")
    rid = getattr(m, "audit_report_id", "")
    old = deps.report_repo.get(rid) if rid else None
    if old is None:
        raise HTTPException(400, "该物料还没有可复用的审核报告,请先完成一次审核")
    svc = deps.get_audit_service()
    pid = getattr(m, "project_id", "") or ""
    job = svc.submit(m.type, oss_key=m.oss_key, owner_id=m.owner_id, material_id=mid,
                     video_kind=("work" if pid else "material"), project_id=pid)
    report = svc.recheck(job, old)                 # 同步重判 + 持久化(_persist 回写 m.audit_status/report_id)
    _sync_task_after_recheck(mid, report)          # 关联任务(若有)同步
    return {"id": mid, "audit_status": report.verdict, "report": _report_out(report)}


@router.get("/audit/queue")
def audit_queue(page: int = Query(1, ge=1), size: int = Query(50, ge=1, le=100),
                type: str | None = None, project: str | None = None, user: dict = Depends(_user)):
    """人工审核队列(管理员):待复核物料 + 可内联播放的签名 URL + 命中原因报告,一次拉齐 → 卡片内直接看直接判。
    project 缺省/"" → 物料栏(无项目);project=P → 项目 P 的待审作品。"""
    _require_perm(user, "materials.audit")
    off, lim = _page_args(page, size)
    items, total = deps.get_library_service().all(
        status=_check_status("review"), type=_check_type(type), project_id=(project or ""),
        offset=off, limit=lim)
    out = []
    for m in items:
        rid = getattr(m, "audit_report_id", "")
        rep = deps.report_repo.get(rid) if rid else None
        rep_out = _report_out(rep) if rep else None
        if rep_out and m.type != MaterialType.CORPUS:
            rep_out = {**rep_out, "segments": []}  # 卡片只用 triggered 命中项;非文本无需回传整条转写(省带宽)
        out.append({**_mat_out(m, uid=user["id"]), "media_url": _media_url(m), "report": rep_out})
    return _page_out(out, total, page, size)


@router.get("/audit/queue/tabs")
def audit_queue_tabs(user: dict = Depends(_user)):
    """审核栏 tab:物料(无项目)+ 每个项目一个 tab,各带待审数量角标。"""
    _require_perm(user, "materials.audit")
    items, _ = deps.get_library_service().all(status=_check_status("review"), limit=None)
    counts: dict[str, int] = {}
    for m in items:
        counts[getattr(m, "project_id", "") or ""] = counts.get(getattr(m, "project_id", "") or "", 0) + 1
    # 项目(作品)在前、物料栏放最后 —— 与上传页「作品/物料」一致的项目优先顺序
    tabs = [{"key": p.id, "label": p.name, "count": counts.get(p.id, 0)} for p in deps.project_repo.list()]
    tabs.append({"key": "", "label": "物料", "count": counts.get("", 0)})
    return {"tabs": tabs}


# ── 语义搜索(F3)——在公共库范围内搜索 ──
@router.get("/search")
def search(q: str = "", page: int = Query(1, ge=1), size: int = Query(24, ge=1, le=100),
           type: str | None = None, tag: str | None = None, user: dict = Depends(_user)):
    off, lim = _page_args(page, size)
    results, total = deps.get_search_service().search(
        q, type=_check_type(type), tag=tag or None, offset=off, limit=lim)
    fav = deps.favorites.material_ids(user["id"])
    return _page_out([_mat_out(m, fav, user["id"]) for m in results], total, page, size, key="results")


# ── 物料库:我的 / 公共 / 全部(管理员)──
@router.get("/library/mine")
def my_library(page: int = Query(1, ge=1), size: int = Query(24, ge=1, le=100),
               type: str | None = None, tag: str | None = None, q: str | None = None,
               project: str | None = None, user: dict = Depends(_user)):
    _require_auth(user)
    off, lim = _page_args(page, size)
    fav = deps.favorites.material_ids(user["id"])
    items, total = deps.get_library_service().mine(
        user["id"], type=_check_type(type), tag=tag or None, keyword=q or None,
        project_id=project, offset=off, limit=lim)
    return _page_out([_mat_out(m, fav, user["id"]) for m in items], total, page, size)


@router.get("/library/public")
def public_library(page: int = Query(1, ge=1), size: int = Query(24, ge=1, le=100),
                   type: str | None = None, tag: str | None = None, q: str | None = None,
                   project: str | None = None, user: dict = Depends(_user)):
    off, lim = _page_args(page, size)
    fav = deps.favorites.material_ids(user["id"])
    items, total = deps.get_library_service().public(
        type=_check_type(type), tag=tag or None, keyword=q or None, project_id=project,
        offset=off, limit=lim)
    return _page_out([_mat_out(m, fav, user["id"]) for m in items], total, page, size)


@router.get("/library/all")
def all_library(page: int = Query(1, ge=1), size: int = Query(24, ge=1, le=100),
                status: str | None = None, type: str | None = None, tag: str | None = None,
                q: str | None = None, project: str | None = None, user: dict = Depends(_user)):
    """管理员:看所有用户的物料。服务端分页/筛选。"""
    _require_perm(user, "library.all")
    off, lim = _page_args(page, size)
    items, total = deps.get_library_service().all(
        status=_check_status(status), type=_check_type(type), tag=tag or None,
        keyword=q or None, project_id=project, offset=off, limit=lim)
    return _page_out([_mat_out(m, uid=user["id"]) for m in items], total, page, size)


@router.post("/materials/{mid}/publish")
def publish(mid: str, user: dict = Depends(_user)):
    """管理员:把物料发布到公共物料库。"""
    _require_perm(user, "materials.publish")
    m = deps.get_library_service().publish(mid, True)
    if m is None:
        raise HTTPException(404, "material not found")
    return _mat_out(m, uid=user["id"])


@router.delete("/materials/{mid}/publish")
def unpublish(mid: str, user: dict = Depends(_user)):
    """管理员:把物料撤出公共物料库。"""
    _require_perm(user, "materials.publish")
    m = deps.get_library_service().publish(mid, False)
    if m is None:
        raise HTTPException(404, "material not found")
    return _mat_out(m, uid=user["id"])


@router.post("/materials/{mid}/favorite")
def favorite(mid: str, user: dict = Depends(_user)):
    """收藏公共物料到我的物料库。仅能收藏公共库(已发布且过审)的物料。"""
    _require_auth(user)
    m = deps.material_repo.get(mid)
    if m is None:
        raise HTTPException(404, "material not found")
    if not (m.is_public and m.audit_status == AuditStatus.PASS):
        raise HTTPException(403, "只能收藏公共物料库中的物料")
    deps.get_library_service().favorite(user["id"], mid)
    return {"favorited": mid}


@router.delete("/materials/{mid}/favorite")
def unfavorite(mid: str, user: dict = Depends(_user)):
    _require_auth(user)
    deps.get_library_service().unfavorite(user["id"], mid)
    return {"unfavorited": mid}


# ── 用户(F7)──
@router.post("/users/register")
def register(body: schemas.RegisterIn):
    try:
        u = deps.get_user_service().register(body.name, body.password)
    except DuplicateName:
        raise HTTPException(409, "用户名已被占用")
    except InvalidCredentials:
        raise HTTPException(400, "用户名和密码不能为空")
    return {"id": u.id, "name": u.name, "role": u.role}


@router.post("/users/login")
def login(body: schemas.LoginIn):
    try:
        token = deps.get_user_service().login(body.name, body.password)
    except InvalidCredentials:
        raise HTTPException(401, "invalid credentials")
    u = deps.user_repo.get_by_name(body.name)
    refresh_token = deps.refresh_token_issuer.issue(u.id)
    return {"token": token, "refreshToken": refresh_token,
            "user": {"id": u.id, "name": u.name, "role": u.role}}


@router.post("/users/refresh")
def refresh_login(body: schemas.RefreshTokenIn):
    uid = deps.refresh_token_issuer.consume(body.refreshToken)
    if uid is None:
        raise HTTPException(401, "refresh token invalid or expired")
    u = deps.user_repo.get(uid)
    if u is None:
        raise HTTPException(401, "refresh token invalid or expired")
    return {
        "token": deps.token_issuer.issue(uid),
        "refreshToken": deps.refresh_token_issuer.issue(uid),
        "user": {"id": u.id, "name": u.name, "role": u.role},
    }


# ── 功能权限后台(F8)──
# 权限目录:每条权限写清楚是什么、什么意思(功能权限页授权弹窗用)
PERM_CATALOG = [
    {"key": "materials.audit", "label": "内容审核复核", "desc": "复核待定物料、操作审核队列(通过/拦截)"},
    {"key": "materials.publish", "label": "发布物料", "desc": "把物料发布到公共库,或从公共库下架"},
    {"key": "library.all", "label": "查看全部物料", "desc": "查看所有用户的物料(管理视图)"},
    {"key": "audit.rules", "label": "审核规则", "desc": "新增 / 删除违规判定规则"},
    {"key": "admin.grant", "label": "权限与账号管理", "desc": "给用户授权 / 收回权限、增删账号"},
    {"key": "materials.delete_any", "label": "删除任意物料", "desc": "删除其他用户上传的物料"},
]
_PERM_KEYS = {p["key"] for p in PERM_CATALOG}


@router.get("/admin/perm-catalog")
def perm_catalog(user: dict = Depends(_user)):
    """可授予的功能权限清单(带中文名+说明),给授权弹窗用。"""
    _require_perm(user, "admin.grant")
    return {"catalog": PERM_CATALOG}


@router.get("/admin/users")
def list_users(page: int = Query(1, ge=1), size: int = Query(20, ge=1, le=100),
               q: str = Query(""), role: str = Query(""), user: dict = Depends(_user)):
    """账号列表(名字 / 角色 / 已被单独授予的权限)。"""
    _require_perm(user, "admin.grant")
    role = (role or "").strip()
    if role and role not in ("admin", "user"):
        raise HTTPException(400, "非法用户角色")
    keyword = (q or "").strip()
    total = deps.user_repo.count(q=keyword, role=role)
    off, lim = _page_args(page, size)
    accounts = deps.user_repo.list(q=keyword, role=role, offset=off, limit=lim)
    permission_map = deps.rbac.user_permissions_for({account.id for account in accounts})
    users = [{"id": account.id, "name": account.name, "role": account.role,
              "permissions": sorted(permission_map.get(account.id, set()))}
             for account in accounts]
    return _page_out(users, total, page, size, key="users")


@router.get("/admin/users/{uid}")
def get_admin_user(uid: str, user: dict = Depends(_user)):
    _require_perm(user, "admin.grant")
    account = deps.user_repo.get(uid)
    if account is None:
        raise HTTPException(404, "用户不存在")
    return {"id": account.id, "name": account.name, "role": account.role,
            "permissions": sorted(deps.rbac.user_permissions(account.id))}


@router.post("/admin/users")
def create_user(body: schemas.UserCreate, user: dict = Depends(_user)):
    """新增账号(默认普通用户;admin 是唯一管理员)。"""
    _require_perm(user, "admin.grant")
    try:
        u = deps.get_user_service().register(body.name, body.password)
    except DuplicateName:
        raise HTTPException(409, "用户名已被占用")
    except InvalidCredentials:
        raise HTTPException(400, "用户名和密码不能为空")
    return {"id": u.id, "name": u.name, "role": u.role}


@router.delete("/admin/users/{uid}")
def delete_user(uid: str, user: dict = Depends(_user)):
    """删除账号。不能删管理员、不能删自己。"""
    _require_perm(user, "admin.grant")
    target = deps.user_repo.get(uid)
    if target is None:
        return {"deleted": uid}
    if target.role == "admin":
        raise HTTPException(400, "不能删除管理员账号")
    if uid == user["id"]:
        raise HTTPException(400, "不能删除自己")
    deps.user_repo.delete(uid)
    return {"deleted": uid}


@router.post("/admin/users/{uid}/perms")
def set_user_perms(uid: str, body: schemas.UserPermsIn, user: dict = Depends(_user)):
    """给某用户设置功能权限(整套替换;只接受权限目录内的权限)。授权即时生效。"""
    _require_perm(user, "admin.grant")
    target = deps.user_repo.get(uid)
    if target is None:
        raise HTTPException(404, "用户不存在")
    perms = {p for p in body.permissions if p in _PERM_KEYS}
    deps.rbac.set_user_permissions(uid, perms)
    return {"id": uid, "name": target.name, "permissions": sorted(perms)}


# 旧的按角色授权端点(保留兼容,前端已改用按用户授权)
@router.post("/admin/grant")
def grant(body: schemas.GrantIn, user: dict = Depends(_user)):
    _require_perm(user, "admin.grant")
    deps.get_authz_service().grant(body.role, body.permission)
    return {"role": body.role, "permissions": sorted(deps.rbac.permissions_of(body.role))}


@router.get("/admin/permissions")
def role_permissions(role: str, user: dict = Depends(_user)):
    _require_perm(user, "admin.grant")
    return {"role": role, "permissions": sorted(deps.rbac.permissions_of(role))}


# ── 内容安全白名单(治误伤:命中这些词即便阿里云判违规也放行)──
@router.get("/admin/whitelist")
def list_whitelist(user: dict = Depends(_user)):
    _require_perm(user, "admin.grant")
    return {"words": deps.whitelist_repo.list()}


@router.post("/admin/whitelist")
def add_whitelist(body: schemas.WhitelistIn, user: dict = Depends(_user)):
    _require_perm(user, "admin.grant")
    for w in body.words:
        deps.whitelist_repo.add(w)
    return {"words": deps.whitelist_repo.list()}


@router.delete("/admin/whitelist")
def remove_whitelist(word: str, user: dict = Depends(_user)):
    _require_perm(user, "admin.grant")
    deps.whitelist_repo.remove(word)
    return {"words": deps.whitelist_repo.list()}


# ── 绝对禁词(审核第一波:命中即拦。管理员精选、非常确定不能讲的硬词)──
@router.get("/admin/blockwords")
def list_blockwords(user: dict = Depends(_user)):
    _require_perm(user, "audit.rules")
    return {"words": deps.blockword_repo.list()}


@router.post("/admin/blockwords")
def add_blockwords(body: schemas.BlockwordIn, user: dict = Depends(_user)):
    _require_perm(user, "audit.rules")
    for w in body.words:
        deps.blockword_repo.add(w)
    return {"words": deps.blockword_repo.list()}


@router.delete("/admin/blockwords")
def remove_blockword(word: str, user: dict = Depends(_user)):
    _require_perm(user, "audit.rules")
    deps.blockword_repo.remove(word)
    return {"words": deps.blockword_repo.list()}


# ── 作品项目(管理员建/删;所有登录用户可列出——供提交选项目 + 浏览筛选)──
@router.get("/projects")
def list_projects(user: dict = Depends(_user)):
    """项目列表(供提交作品选项目 + 分项目浏览/审核栏)。登录即可读。
    自愈:为空时自动补默认项目 → 作品的项目下拉永不为空,作品不会因「没项目可选」而上传失败。"""
    _require_auth(user)
    deps.ensure_default_project()
    return {"projects": [_project_out(p) for p in deps.project_repo.list()]}


@router.post("/admin/projects")
def add_project(body: schemas.ProjectIn, user: dict = Depends(_user)):
    """新建作品项目。管理项目/规则复用 audit.rules 权限。名字不能空/重复。"""
    _require_perm(user, "audit.rules")
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(400, "项目名不能为空。")
    if deps.project_repo.get_by_name(name) is not None:
        raise HTTPException(409, "项目名已存在。")
    p = Project(id=next_id_str(), name=name, created_by=user["id"], created_ms=int(time.time() * 1000))
    deps.project_repo.add(p)
    return _project_out(p)


@router.delete("/admin/projects/{project_id}")
def delete_project(project_id: str, user: dict = Depends(_user)):
    """删除项目:项目下还有作品(Material.project_id==id)则禁止删除;否则连带删它的规则。"""
    _require_perm(user, "audit.rules")
    if deps.project_repo.get(project_id) is None:
        return {"deleted": project_id}
    if len(deps.project_repo.list()) <= 1:                 # 守最后一个:作品必须有项目可归属,不能删到零
        raise HTTPException(400, "至少保留一个项目;作品需要归属项目。")
    _, n = deps.get_library_service().all(project_id=project_id, limit=0)
    if n > 0:
        raise HTTPException(400, f"该项目下还有 {n} 个作品,请先处理这些作品再删除项目。")
    for r in deps.rule_repo.list():                        # 连带删该项目的规则
        if getattr(r, "project_id", "") == project_id:
            deps.rule_repo.delete(r.id, by=user["id"])
    deps.project_repo.delete(project_id)
    return {"deleted": project_id}


# ── 上传账号 / 素材提报(管理员) ──
@router.post("/admin/uploads/file")
async def upload_admin_file(file: UploadFile = File(...), scope: str = Form("uploads"),
                            submission_id: str = Form(""), upload_id: str = Form(""),
                            user: dict = Depends(_user)):
    _require_auth(user)
    safe_upload_id = (upload_id or "").strip().lower()
    if safe_upload_id and not _UPLOAD_ID_RE.fullmatch(safe_upload_id):
        raise HTTPException(400, "上传任务 ID 格式不正确")
    safe_name = _safe_upload_file_name(file.filename or "")
    safe_scope = _safe_upload_scope(scope)
    if safe_scope == "short-drama-tasks":
        suffix = Path(safe_name).suffix.lower()
        allowed_image_suffixes = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}
        if suffix not in allowed_image_suffixes or not (file.content_type or "").lower().startswith("image/"):
            raise HTTPException(400, "短剧任务附件仅支持 JPG、PNG、GIF、WebP 或 BMP 图片。")
    if user.get("role") != "admin":
        if safe_scope in ("submissions", "requirements", "templates", "short-drama-tasks"):
            pass
        elif safe_scope == "submissions/rejects" and submission_id:
            _require_submission_access(user, submission_id, edit=True)
        else:
            raise HTTPException(403, "无权向该目录上传文件")
    key = f"{safe_scope}/{uuid.uuid4().hex}-{safe_name}"
    await file.seek(0)
    total_size = _upload_file_size(file.file)
    if safe_scope == "short-drama-tasks" and total_size > 10 * 1024 * 1024:
        raise HTTPException(400, "短剧任务图片不能超过 10 MB。")
    if safe_upload_id:
        _set_upload_progress(
            safe_upload_id, user["id"], status="uploading", stage="storage",
            loaded=0, total=total_size, file_name=safe_name,
        )

    def report_progress(consumed: int, total: int) -> None:
        measured_total = int(total or total_size or 0)
        _set_upload_progress(
            safe_upload_id, user["id"], status="uploading", stage="storage",
            loaded=max(0, int(consumed or 0)), total=max(0, measured_total),
            file_name=safe_name,
        )

    try:
        if safe_upload_id:
            await run_in_threadpool(deps.storage.put_fileobj, key, file.file, report_progress)
        else:
            # 保留原调用形式，兼容现有 Skill 和只实现双参数方法的调用方。
            await run_in_threadpool(deps.storage.put_fileobj, key, file.file)
    except Exception as exc:
        _set_upload_progress(
            safe_upload_id, user["id"], status="error", stage="storage",
            error="附件上传失败，请检查 OSS 网络连接或配置",
        )
        # 云存储网络/配置异常不应裸露成无信息的 500。
        import logging
        logging.getLogger(__name__).exception("附件上传到 OSS 失败: key=%s", key)
        raise HTTPException(502, "附件上传失败，请检查 OSS 网络连接或配置") from exc
    decoded_oss_key = ""
    requires_decode = 0
    if safe_scope == "submissions" and os.path.splitext(safe_name)[1].lower() in (".mp4", ".mov", ".m4v"):
        decoded_file = None
        try:
            _set_upload_progress(
                safe_upload_id, user["id"], status="processing", stage="inspecting",
                loaded=0, total=100, progress=0, file_name=safe_name,
            )

            def report_transcode_progress(percent: int) -> None:
                normalized = max(0, min(100, int(percent or 0)))
                _set_upload_progress(
                    safe_upload_id, user["id"], status="processing", stage="transcoding",
                    loaded=normalized, total=100, progress=normalized, file_name=safe_name,
                )

            decoded_file = await run_in_threadpool(
                transcode_hevc_to_h264, file.file, report_transcode_progress,
            )
            if decoded_file is not None:
                requires_decode = 1
                stem = os.path.splitext(safe_name)[0] or "video"
                decoded_oss_key = f"{safe_scope}/decoded/{uuid.uuid4().hex}-{stem}-h264.mp4"
                decoded_size = _upload_file_size(decoded_file)

                def report_decoded_upload(consumed: int, total: int) -> None:
                    measured_total = int(total or decoded_size or 0)
                    _set_upload_progress(
                        safe_upload_id, user["id"], status="processing", stage="decoded_storage",
                        loaded=max(0, int(consumed or 0)), total=max(0, measured_total),
                        progress=100, file_name=safe_name,
                    )

                if safe_upload_id:
                    await run_in_threadpool(
                        deps.storage.put_fileobj, decoded_oss_key, decoded_file, report_decoded_upload,
                    )
                else:
                    await run_in_threadpool(deps.storage.put_fileobj, decoded_oss_key, decoded_file)
        except Exception:
            import logging
            logging.getLogger(__name__).exception(
                "HEVC 兼容版生成失败，保留源文件: key=%s", key,
            )
            if decoded_oss_key:
                try:
                    await run_in_threadpool(deps.storage.delete, decoded_oss_key)
                except Exception:
                    pass
            decoded_oss_key = ""
        finally:
            if decoded_file is not None:
                decoded_file.close()

    _set_upload_progress(
        safe_upload_id, user["id"], status="done", stage="storage",
        loaded=total_size, total=total_size, file_name=safe_name,
    )
    return {
        "oss_key": key,
        "decoded_oss_key": decoded_oss_key,
        "requires_decode": requires_decode,
        "file_name": safe_name,
    }


@router.get("/admin/uploads/progress/{upload_id}")
def admin_upload_progress(upload_id: str, user: dict = Depends(_user)):
    _require_auth(user)
    safe_upload_id = (upload_id or "").strip().lower()
    if not _UPLOAD_ID_RE.fullmatch(safe_upload_id):
        raise HTTPException(400, "上传任务 ID 格式不正确")
    with _UPLOAD_PROGRESS_LOCK:
        item = dict(_UPLOAD_PROGRESS.get(safe_upload_id) or {})
    if not item:
        raise HTTPException(404, "上传任务尚未开始")
    if item.get("user_id") != user["id"]:
        raise HTTPException(403, "无权查看该上传任务")
    return {key: value for key, value in item.items() if key not in ("user_id", "updated_at")}


@router.get("/admin/uploads/url")
def admin_upload_signed_url(key: str = Query(...), dl: int = Query(0),
                            submission_id: str = Query(""), requirement_id: str = Query(""),
                            template_id: str = Query(""), short_drama_task_id: str = Query(""),
                            recycle_bin: bool = Query(False),
                            user: dict = Depends(_user)):
    _require_auth(user)
    if not key or not key.strip():
        raise HTTPException(400, "缺少 oss key")
    k = key.strip()
    if user.get("role") != "admin":
        if template_id:
            template = _require_template_access(user, template_id)
            if k not in {template.reference_oss_key, template.bgm_oss_key}:
                raise HTTPException(403, "No permission to access this template asset")
        elif short_drama_task_id:
            task = _require_short_drama_task(short_drama_task_id, user)
            if k not in {task.cover_oss_key, task.data_image_oss_key}:
                raise HTTPException(403, "无权访问该短剧任务图片")
        elif requirement_id:
            requirement = _require_requirement(requirement_id, user)
            if k not in set(requirement.attachments or []):
                raise HTTPException(403, "无权访问该文件")
        elif not submission_id:
            raise HTTPException(403, "缺少素材提报权限上下文")
        else:
            submission = (_require_deleted_submission_access(user, submission_id)
                          if recycle_bin else _require_submission_access(user, submission_id))
            allowed_keys = {
                submission.oss_key, submission.decoded_oss_key,
                *(submission.platform_reject_attachments or []),
            }
            allowed_keys.update(_submission_operation_asset_keys(submission_id))
            if k not in allowed_keys:
                raise HTTPException(403, "无权访问该文件")
    try:
        url = deps.storage.download_url(k) if dl == 1 else deps.storage.signed_url(k)
    except Exception:
        raise HTTPException(404, "获取签名 URL 失败")
    return {"url": url}


@router.get("/admin/video-editing-templates")
def list_video_editing_templates(name: str = Query(""), status: str = Query(""),
                                 page: int = Query(1, ge=1), size: int = Query(20, ge=1, le=100),
                                 user: dict = Depends(_user)):
    _require_auth(user)
    checked_status = _template_status(status) if status else ""
    clean_name = (name or "").strip()
    offset = (page - 1) * size
    items = deps.video_editing_template_repo.list(
        name=clean_name, status=checked_status, offset=offset, limit=size,
    )
    return {
        "templates": [_template_out(item, user) for item in items],
        "count": deps.video_editing_template_repo.count(name=clean_name, status=checked_status),
        "page": page,
        "size": size,
    }


@router.get("/admin/video-editing-templates/by-name/{template_name}")
def get_video_editing_template_by_name(template_name: str, user: dict = Depends(_user)):
    _require_auth(user)
    template = deps.video_editing_template_repo.get_by_name(template_name)
    if template is None:
        raise HTTPException(404, "Video-editing template not found")
    return _template_out(template, user)


@router.get("/admin/video-editing-templates/{template_id}")
def get_video_editing_template(template_id: str, user: dict = Depends(_user)):
    return _template_out(_require_template_access(user, template_id), user)


@router.post("/admin/video-editing-templates")
def create_video_editing_template(body: schemas.VideoEditingTemplateIn, user: dict = Depends(_user)):
    _require_auth(user)
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(400, "Template name is required")
    if deps.video_editing_template_repo.get_by_name(name) is not None:
        raise HTTPException(409, "Template name already exists")
    template = VideoEditingTemplate(
        id=next_id_str(), name=name, description=(body.description or "").strip(),
        reference_oss_key=_template_oss_key(body.reference_oss_key, "reference_oss_key"),
        narration_voice=dict(body.narration_voice or {}),
        bgm_oss_key=_template_oss_key(body.bgm_oss_key, "BGM"),
        config=dict(body.config or {}), status=_template_status(body.status),
        version=1, created_by=user["id"],
    )
    try:
        deps.video_editing_template_repo.save(template, by=user["id"])
    except ValueError as exc:
        raise HTTPException(409, str(exc)) from exc
    return _template_out(deps.video_editing_template_repo.get(template.id) or template, user)


@router.put("/admin/video-editing-templates/{template_id}")
def update_video_editing_template(template_id: str, body: schemas.VideoEditingTemplateUpdateIn,
                                  user: dict = Depends(_user)):
    template = _require_template_access(user, template_id, edit=True)
    fields_set = getattr(body, "model_fields_set", None)
    if fields_set is None:
        fields_set = getattr(body, "__fields_set__", set())
    if not fields_set:
        raise HTTPException(400, "At least one template field is required")
    if body.name is not None:
        name = body.name.strip()
        if not name:
            raise HTTPException(400, "Template name is required")
        duplicate = deps.video_editing_template_repo.get_by_name(name)
        if duplicate is not None and duplicate.id != template.id:
            raise HTTPException(409, "Template name already exists")
        template.name = name
    if body.description is not None:
        template.description = body.description.strip()
    if body.reference_oss_key is not None:
        template.reference_oss_key = _template_oss_key(body.reference_oss_key, "reference_oss_key")
    if body.narration_voice is not None:
        template.narration_voice = dict(body.narration_voice)
    if body.bgm_oss_key is not None:
        template.bgm_oss_key = _template_oss_key(body.bgm_oss_key, "BGM")
    if body.config is not None:
        template.config = dict(body.config)
    if body.status is not None:
        template.status = _template_status(body.status)
    template.version += 1
    try:
        deps.video_editing_template_repo.save(template, by=user["id"])
    except ValueError as exc:
        raise HTTPException(409, str(exc)) from exc
    return _template_out(deps.video_editing_template_repo.get(template.id) or template, user)


_SHORT_DRAMA_TASK_STATUSES = {"未上线", "已上线", "已结束"}
_SHORT_DRAMA_UPLOAD_DEGREES = {"无人问津", "适中", "饱和"}
_SHORT_DRAMA_TASK_SORT_FIELDS = {
    "online_time", "expiration_time", "drama_name", "task_type", "theme", "task_status", "task_id",
    "settlement_mode", "settlement_period", "created_time",
}
_SHORT_DRAMA_TASK_SORT_ORDERS = {"asc", "desc"}
_SHORT_DRAMA_TASK_EXCEL_COLUMNS = [
    ("剧名", "drama_name"), ("任务ID", "task_id"), ("类型", "task_type"),
    ("题材", "theme"), ("标签", "tags"), ("任务上线时间", "online_time"),
    ("任务到期时间", "expiration_time"),
    ("任务状态", "task_status"), ("要求", "requirements"),
    ("话题/剪辑要求", "topic_editing_requirements"),
    ("网盘素材", "cloud_material_url"), ("优质案例", "quality_case"),
    ("封面", "cover_oss_key"), ("备注", "remarks"),
]


def _empty_short_drama_upload_summary() -> dict:
    return {
        "team_names": [], "upload_count": 0,
        "can_upload_count": 0, "publish_success_count": 0,
    }


def _short_drama_upload_summaries(items: list[ShortDramaTask]) -> dict[str, dict]:
    if not items:
        return {}
    return deps.material_submission_repo.aggregate_uploads_by_drama_names(
        [(item.id, item.drama_name) for item in items]
    )


def _short_drama_upload_degree(item: ShortDramaTask,
                               upload_summary: dict | None = None) -> tuple[str, int]:
    teams = {
        str(name or "").strip().casefold()
        for name in [
            *(item.pre_upload_teams or []),
            *((upload_summary or {}).get("team_names") or []),
        ]
        if str(name or "").strip()
    }
    team_count = len(teams)
    if team_count == 0:
        return "无人问津", team_count
    if team_count >= 3:
        return "饱和", team_count
    return "适中", team_count


def _short_drama_task_out(item: ShortDramaTask, upload_summary: dict | None = None) -> dict:
    summary = dict(upload_summary or _empty_short_drama_upload_summary())
    upload_degree, upload_degree_team_count = _short_drama_upload_degree(item, summary)
    summary.update(upload_degree=upload_degree,
                   upload_degree_team_count=upload_degree_team_count)
    return {
        "id": item.id, "online_time": item.online_time,
        "expiration_time": item.expiration_time, "drama_name": item.drama_name,
        "task_type": item.task_type, "tags": list(item.tags or []),
        "pre_upload_teams": list(item.pre_upload_teams or []), "theme": item.theme,
        "task_status": item.task_status, "task_id": item.task_id,
        "requirements": item.requirements, "cover_oss_key": item.cover_oss_key,
        "cloud_material_url": item.cloud_material_url,
        "topic_editing_requirements": item.topic_editing_requirements,
        "submission_activity_time": item.submission_activity_time,
        "settlement_mode": item.settlement_mode,
        "commission_validity_period": item.commission_validity_period,
        "settlement_period": item.settlement_period,
        "data_image_oss_key": item.data_image_oss_key, "quality_case": item.quality_case,
        "remarks": item.remarks, "created_by": item.created_by,
        "created_time": item.created_time, "updated_by": item.updated_by,
        "updated_time": item.updated_time,
        "upload_summary": summary, "upload_degree": upload_degree,
        "upload_degree_team_count": upload_degree_team_count,
    }


def _short_drama_task_from_body(body: schemas.ShortDramaTaskIn, *, item_id: str,
                                created_by: str) -> ShortDramaTask:
    drama_name = (body.drama_name or "").strip()
    if not drama_name:
        raise HTTPException(400, "剧名不能为空")
    cloud_material_url = (body.cloud_material_url or "").strip()
    if not cloud_material_url:
        raise HTTPException(400, "网盘素材不能为空")
    cover_oss_key = (body.cover_oss_key or "").strip()
    status = (body.task_status or "未上线").strip()
    if status not in _SHORT_DRAMA_TASK_STATUSES:
        raise HTTPException(400, "任务状态只能是未上线、已上线或已结束")
    tags: list[str] = []
    for value in body.tags or []:
        tag = str(value or "").strip()
        if tag and tag not in tags:
            tags.append(tag)
    pre_upload_teams: list[str] = []
    seen_pre_upload_teams: set[str] = set()
    for value in body.pre_upload_teams or []:
        team_name = str(value or "").strip()
        if len(team_name) > 500:
            raise HTTPException(400, "团队名称不能超过 500 个字符")
        normalized_team_name = team_name.casefold()
        if team_name and normalized_team_name not in seen_pre_upload_teams:
            seen_pre_upload_teams.add(normalized_team_name)
            pre_upload_teams.append(team_name)
    return ShortDramaTask(
        id=item_id, online_time=(body.online_time or "").strip(),
        expiration_time=(body.expiration_time or "").strip(), drama_name=drama_name,
        task_type=(body.task_type or "").strip(), tags=tags,
        pre_upload_teams=pre_upload_teams, theme=(body.theme or "").strip(),
        task_status=status, task_id=(body.task_id or "").strip(),
        requirements=(body.requirements or "").strip(),
        cover_oss_key=cover_oss_key,
        cloud_material_url=cloud_material_url,
        topic_editing_requirements=(body.topic_editing_requirements or "").strip(),
        submission_activity_time=(body.submission_activity_time or "").strip(),
        settlement_mode=(body.settlement_mode or "").strip(),
        commission_validity_period=(body.commission_validity_period or "").strip(),
        settlement_period=(body.settlement_period or "").strip(),
        data_image_oss_key=(body.data_image_oss_key or "").strip(),
        quality_case=(body.quality_case or "").strip(), remarks=(body.remarks or "").strip(),
        created_by=created_by,
    )


def _require_short_drama_task(item_id: str, user: dict) -> ShortDramaTask:
    _require_auth(user)
    item = deps.short_drama_task_repo.get(item_id)
    if not item:
        raise HTTPException(404, "短剧任务不存在")
    return item


def _short_drama_task_sort(sort_by: str, sort_order: str) -> tuple[str, str]:
    checked_by = (sort_by or "created_time").strip()
    checked_order = (sort_order or "desc").strip().lower()
    if checked_by not in _SHORT_DRAMA_TASK_SORT_FIELDS:
        raise HTTPException(400, "非法短剧任务排序字段")
    if checked_order not in _SHORT_DRAMA_TASK_SORT_ORDERS:
        raise HTTPException(400, "非法短剧任务排序方向")
    return checked_by, checked_order


@router.get("/admin/short-drama-tasks")
def list_short_drama_tasks(
    drama_name: str = "", task_status: str = "", task_type: str = "", theme: str = "",
    online_time: str = "", expiration_time: str = "", task_id: str = "", tag: str = "", pre_upload_team: str = "",
    actual_upload_team: str = "", upload_degree: str = "", remarks: str = "",
    sort_by: str = "created_time", sort_order: str = "desc",
    page: int = Query(1, ge=1), size: int = Query(10, ge=1, le=100),
    user: dict = Depends(_user),
):
    _require_auth(user)
    if task_status and task_status not in _SHORT_DRAMA_TASK_STATUSES:
        raise HTTPException(400, "非法任务状态")
    checked_upload_degree = upload_degree.strip()
    if checked_upload_degree and checked_upload_degree not in _SHORT_DRAMA_UPLOAD_DEGREES:
        raise HTTPException(400, "非法上传程度")
    offset, limit = _page_args(page, size)
    actual_team = actual_upload_team.strip()
    actual_upload_drama_names = (
        deps.material_submission_repo.list_drama_names_by_team(actual_team)
        if actual_team else None
    )
    filters = dict(drama_name=drama_name.strip(), task_status=task_status.strip(),
                   task_type=task_type.strip(), theme=theme.strip(),
                   online_time=online_time.strip(), expiration_time=expiration_time.strip(),
                   task_id=task_id.strip(), remarks=remarks.strip(),
                   tag=tag.strip(), pre_upload_team=pre_upload_team.strip(),
                   actual_upload_drama_names=actual_upload_drama_names)
    checked_sort_by, checked_sort_order = _short_drama_task_sort(sort_by, sort_order)
    if checked_upload_degree:
        filtered_items = deps.short_drama_task_repo.list(
            offset=0, limit=None, sort_by=checked_sort_by,
            sort_order=checked_sort_order, **filters,
        )
        all_summaries = _short_drama_upload_summaries(filtered_items)
        matched_items = [
            item for item in filtered_items
            if _short_drama_upload_degree(item, all_summaries.get(item.id))[0]
            == checked_upload_degree
        ]
        total = len(matched_items)
        items = matched_items[offset:offset + limit]
        summaries = {item.id: all_summaries.get(item.id) for item in items}
    else:
        total = deps.short_drama_task_repo.count(**filters)
        items = deps.short_drama_task_repo.list(
            offset=offset, limit=limit, sort_by=checked_sort_by,
            sort_order=checked_sort_order, **filters,
        )
        summaries = _short_drama_upload_summaries(items)
    return _page_out(
        [_short_drama_task_out(x, summaries.get(x.id)) for x in items],
        total, page, size, key="tasks",
    )


@router.get("/admin/short-drama-tasks/options/{field}")
def list_short_drama_task_options(
    field: str, keyword: str = Query(""), q: str = Query(""),
    limit: int = Query(200, ge=1, le=1000), user: dict = Depends(_user),
):
    _require_auth(user)
    try:
        items = deps.short_drama_task_repo.list_options(
            field, keyword=(keyword or q).strip(), limit=limit,
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    return {"items": items}


@router.get("/admin/short-drama-tasks/import-template")
def download_short_drama_task_import_template(user: dict = Depends(_user)):
    _require_auth(user)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "短剧任务导入"
    sheet.append([x[0] for x in _SHORT_DRAMA_TASK_EXCEL_COLUMNS])
    sheet.append([
        "示例短剧（请删除示例行）", "TASK-001", "短剧", "都市", "爆剧、新剧",
        "2026-08-20", "2026-09-20", "未上线", "示例要求", "示例话题及剪辑要求",
        "https://pan.example.com/example", "https://example.com/case", "", "",
    ])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C85A3E")
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(40, max(14, max(len(str(c.value or "")) for c in column) + 2))
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="short-drama-task-import-template.xlsx"'},
    )


def _excel_tags(value) -> list[str]:
    raw = str(value or "").strip()
    if not raw:
        return []
    if raw.startswith("["):
        try:
            parsed = json.loads(raw)
        except json.JSONDecodeError as exc:
            raise ValueError("标签不是合法 JSON 数组") from exc
        if not isinstance(parsed, list):
            raise ValueError("标签必须是 JSON 数组")
        values = parsed
    else:
        values = re.split(r"[,，、;；\r\n]+", raw)
    return list(dict.fromkeys(str(x or "").strip() for x in values if str(x or "").strip()))


def _validate_public_image_url(url: str) -> str:
    parsed = urlparse((url or "").strip())
    if parsed.scheme not in {"http", "https"} or not parsed.hostname or parsed.username or parsed.password:
        raise ValueError("附件必须填写 http 或 https 图片链接")
    port = parsed.port or (443 if parsed.scheme == "https" else 80)
    try:
        addresses = socket.getaddrinfo(parsed.hostname, port, type=socket.SOCK_STREAM)
    except OSError as exc:
        raise ValueError("附件图片链接无法解析") from exc
    if not addresses or any(
        not ipaddress.ip_address(row[4][0].split("%", 1)[0]).is_global for row in addresses
    ):
        raise ValueError("附件图片链接必须指向公网地址")
    return parsed.geturl()


def _short_task_image_extension(data: bytes) -> str:
    if data.startswith(b"\x89PNG\r\n\x1a\n"):
        return "png"
    if data.startswith(b"\xff\xd8\xff"):
        return "jpg"
    if data.startswith((b"GIF87a", b"GIF89a")):
        return "gif"
    if len(data) >= 12 and data[:4] == b"RIFF" and data[8:12] == b"WEBP":
        return "webp"
    if data.startswith(b"BM"):
        return "bmp"
    raise ValueError("附件链接内容不是支持的图片格式")


async def _download_short_task_image(client, url: str) -> str:
    current = url
    for _ in range(4):
        current = await run_in_threadpool(_validate_public_image_url, current)
        async with client.stream("GET", current) as response:
            if response.status_code in {301, 302, 303, 307, 308}:
                location = response.headers.get("location", "")
                if not location:
                    raise ValueError("附件图片链接跳转地址无效")
                current = urljoin(current, location)
                continue
            if response.status_code < 200 or response.status_code >= 300:
                raise ValueError("附件图片链接下载失败")
            payload = bytearray()
            async for chunk in response.aiter_bytes():
                payload.extend(chunk)
                if len(payload) > 10 * 1024 * 1024:
                    raise ValueError("单张附件图片不能超过 10MB")
        extension = _short_task_image_extension(bytes(payload))
        key = f"short-drama-tasks/{uuid.uuid4().hex}-excel-image.{extension}"
        await run_in_threadpool(deps.storage.put_fileobj, key, io.BytesIO(payload))
        return key
    raise ValueError("附件图片链接跳转次数过多")


async def _store_short_task_excel_images(items: list[ShortDramaTask]) -> None:
    import httpx
    bindings: dict[str, list[tuple[ShortDramaTask, str]]] = {}
    for item in items:
        for field in ("cover_oss_key", "data_image_oss_key"):
            url = str(getattr(item, field, "") or "").strip()
            if url:
                bindings.setdefault(url, []).append((item, field))
    if not bindings:
        return
    semaphore = asyncio.Semaphore(6)

    async def download(client, url):
        async with semaphore:
            return await _download_short_task_image(client, url)

    urls = list(bindings)
    async with httpx.AsyncClient(timeout=20.0, follow_redirects=False) as client:
        results = await asyncio.gather(
            *(download(client, url) for url in urls), return_exceptions=True,
        )
    created_keys = [value for value in results if isinstance(value, str)]
    failure = next((value for value in results if isinstance(value, Exception)), None)
    if failure:
        for key in created_keys:
            try:
                await run_in_threadpool(deps.storage.delete, key)
            except Exception:
                pass
        raise HTTPException(400, str(failure)) from failure
    for url, key in zip(urls, results):
        for item, field in bindings[url]:
            setattr(item, field, key)


@router.post("/admin/short-drama-tasks/import")
async def import_short_drama_tasks(file: UploadFile = File(...), user: dict = Depends(_user)):
    _require_auth(user)
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(400, "仅支持 .xlsx 文件")
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(400, "Excel 文件不能超过 20MB")
    from openpyxl import load_workbook
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(x or "").strip() for x in next(rows)]
    except (StopIteration, ValueError, KeyError, OSError) as exc:
        raise HTTPException(400, "无法读取 Excel 文件") from exc
    expected = {label: key for label, key in _SHORT_DRAMA_TASK_EXCEL_COLUMNS}
    missing = [label for label in ("剧名", "网盘素材") if label not in headers]
    if missing:
        raise HTTPException(400, f"缺少必填列：{missing[0]}")
    column_keys = [expected.get(x, "") for x in headers]
    deduplicated: dict[str, ShortDramaTask] = {}
    errors: list[str] = []
    for row_number, row in enumerate(rows, start=2):
        raw = {key: row[index] for index, key in enumerate(column_keys)
               if key and index < len(row)}
        if not any(value not in (None, "") for value in raw.values()):
            continue
        try:
            payload = {key: str(raw.get(key) or "").strip()
                       for _, key in _SHORT_DRAMA_TASK_EXCEL_COLUMNS if key != "tags"}
            payload["tags"] = _excel_tags(raw.get("tags"))
            for asset_field in ("cover_oss_key",):
                asset_url = payload.get(asset_field, "")
                if asset_url and urlparse(asset_url).scheme not in {"http", "https"}:
                    raise ValueError("封面请填写 http 或 https 图片链接")
            task = _short_drama_task_from_body(
                schemas.ShortDramaTaskIn(**payload), item_id=next_id_str(), created_by=user["id"],
            )
            deduplicated[task.drama_name.casefold()] = task
        except (ValueError, TypeError, HTTPException) as exc:
            detail = exc.detail if isinstance(exc, HTTPException) else str(exc)
            errors.append(f"第 {row_number} 行：{detail}")
        if len(deduplicated) > 5000:
            raise HTTPException(400, "单次最多导入 5000 条任务")
    if errors:
        raise HTTPException(400, {"message": "导入校验失败", "errors": errors[:50]})
    if not deduplicated:
        raise HTTPException(400, "Excel 中没有可导入的数据")
    imported_tasks = list(deduplicated.values())
    await _store_short_task_excel_images(imported_tasks)
    created, updated = deps.short_drama_task_repo.bulk_upsert(imported_tasks, by=user["id"])
    return {"total": len(deduplicated), "created": created, "updated": updated}


@router.get("/admin/short-drama-tasks/{item_id}")
def get_short_drama_task(item_id: str, user: dict = Depends(_user)):
    item = _require_short_drama_task(item_id, user)
    summary = _short_drama_upload_summaries([item]).get(item.id)
    return _short_drama_task_out(item, summary)


@router.put("/admin/short-drama-tasks/{item_id}/pre-upload-teams")
def set_short_drama_task_pre_upload_teams(
    item_id: str, body: schemas.ShortDramaTaskPreUploadTeamsIn,
    user: dict = Depends(_user),
):
    current = _require_short_drama_task(item_id, user)
    team_names: list[str] = []
    seen_team_names: set[str] = set()
    for value in body.team_names or []:
        team_name = str(value or "").strip()
        if len(team_name) > 500:
            raise HTTPException(400, "团队名称不能超过 500 个字符")
        normalized_team_name = team_name.casefold()
        if team_name and normalized_team_name not in seen_team_names:
            seen_team_names.add(normalized_team_name)
            team_names.append(team_name)
    current.pre_upload_teams = team_names
    deps.short_drama_task_repo.save(current, by=user["id"])
    saved = deps.short_drama_task_repo.get(item_id) or current
    summary = _short_drama_upload_summaries([saved]).get(saved.id)
    return _short_drama_task_out(saved, summary)


@router.post("/admin/short-drama-tasks")
def create_short_drama_task(body: schemas.ShortDramaTaskIn, user: dict = Depends(_user)):
    _require_auth(user)
    item = _short_drama_task_from_body(body, item_id=next_id_str(), created_by=user["id"])
    if deps.short_drama_task_repo.get_by_drama_name(item.drama_name):
        raise HTTPException(409, "剧名已存在")
    try:
        deps.short_drama_task_repo.save(item, by=user["id"])
    except ValueError as exc:
        raise HTTPException(409, str(exc)) from exc
    saved = deps.short_drama_task_repo.get(item.id) or item
    summary = _short_drama_upload_summaries([saved]).get(saved.id)
    return _short_drama_task_out(saved, summary)


@router.put("/admin/short-drama-tasks/{item_id}")
def update_short_drama_task(item_id: str, body: schemas.ShortDramaTaskIn,
                            user: dict = Depends(_user)):
    current = _require_short_drama_task(item_id, user)
    item = _short_drama_task_from_body(body, item_id=current.id, created_by=current.created_by)
    item.pre_upload_teams = list(current.pre_upload_teams or [])
    duplicate = deps.short_drama_task_repo.get_by_drama_name(item.drama_name)
    if duplicate and duplicate.id != item.id:
        raise HTTPException(409, "剧名已存在")
    try:
        deps.short_drama_task_repo.save(item, by=user["id"])
    except ValueError as exc:
        raise HTTPException(409, str(exc)) from exc
    saved = deps.short_drama_task_repo.get(item.id) or item
    summary = _short_drama_upload_summaries([saved]).get(saved.id)
    return _short_drama_task_out(saved, summary)


@router.delete("/admin/short-drama-tasks/{item_id}")
def delete_short_drama_task(item_id: str, user: dict = Depends(_user)):
    _require_short_drama_task(item_id, user)
    if not deps.short_drama_task_repo.delete(item_id):
        raise HTTPException(404, "短剧任务不存在")
    return {"ok": True, "id": item_id}


def _visible_material_submissions(user: dict, **filters) -> list[MaterialSubmission]:
    _require_auth(user)
    visible_to_user_id = "" if user.get("role") == "admin" else user.get("id", "")
    return deps.material_submission_repo.list(
        offset=0, limit=None, visible_to_user_id=visible_to_user_id, **filters,
    )


_MATERIAL_SUBMISSION_SORT_FIELDS = {"team_name", "upload_date", "created_time"}
_MATERIAL_SUBMISSION_SORT_ORDERS = {"asc", "desc"}


def _validate_material_submission_sort(sort_by: str = "created_time",
                                       sort_order: str = "desc") -> tuple[str, str]:
    sort_by = (sort_by or "").strip()
    sort_order = (sort_order or "").strip()
    if sort_by and sort_by not in _MATERIAL_SUBMISSION_SORT_FIELDS:
        raise HTTPException(400, "非法素材提报排序字段")
    if sort_by and sort_order not in _MATERIAL_SUBMISSION_SORT_ORDERS:
        raise HTTPException(400, "非法素材提报排序方向")
    if not sort_by and sort_order:
        raise HTTPException(400, "排序方向必须与排序字段同时设置")

    return sort_by, sort_order


def _visible_submission_names(user: dict, field: str, keyword: str, limit: int,
                              recycle_bin: bool = False) -> list[str]:
    key = (keyword or "").lower()
    seen: set[str] = set()
    result: list[str] = []
    for submission in reversed(_visible_material_submissions(user, recycle_bin=recycle_bin)):
        value = str(getattr(submission, field, "") or "").strip()
        if not value or (key and key not in value.lower()) or value in seen:
            continue
        seen.add(value)
        result.append(value)
        if len(result) >= limit:
            break
    return result


@router.get("/admin/material-submissions/upload-account-names")
def list_material_submission_upload_account_names(
    keyword: str = Query(""), q: str = Query(""), limit: int = Query(200, ge=1, le=1000),
    recycle_bin: bool = Query(False),
    user: dict = Depends(_user),
):
    return {"items": _visible_submission_names(
        user, "upload_account_name", (keyword or q).strip(), limit, recycle_bin,
    )}


@router.get("/admin/material-submissions/designated-upload-account-names")
def list_material_submission_designated_upload_account_names(
    keyword: str = Query(""), q: str = Query(""), limit: int = Query(200, ge=1, le=1000),
    recycle_bin: bool = Query(False),
    user: dict = Depends(_user),
):
    return {"items": _visible_submission_names(
        user, "designated_upload_account_name", (keyword or q).strip(), limit, recycle_bin,
    )}


@router.get("/admin/material-submissions/drama-names")
def list_material_submission_drama_names(
    keyword: str = Query(""), q: str = Query(""), limit: int = Query(200, ge=1, le=1000),
    recycle_bin: bool = Query(False),
    user: dict = Depends(_user),
):
    return {"items": _visible_submission_names(
        user, "drama_name", (keyword or q).strip(), limit, recycle_bin,
    )}


@router.get("/admin/material-submissions/team-names")
def list_material_submission_team_names(
    keyword: str = Query(""), q: str = Query(""), limit: int = Query(200, ge=1, le=1000),
    recycle_bin: bool = Query(False),
    user: dict = Depends(_user),
):
    return {"items": _visible_submission_names(
        user, "team_name", (keyword or q).strip(), limit, recycle_bin,
    )}


@router.get("/admin/material-submissions/creator-accounts")
def list_material_submission_creator_accounts(recycle_bin: bool = Query(False),
                                               user: dict = Depends(_user)):
    _require_auth(user)
    optimized_list = getattr(deps.material_submission_repo, "list_creator_accounts", None)
    if callable(optimized_list):
        visible_to_user_id = "" if user.get("role") == "admin" else user.get("id", "")
        return {"items": optimized_list(
            recycle_bin=recycle_bin,
            visible_to_user_id=visible_to_user_id,
        )}
    seen: set[str] = set()
    items = []
    for submission in reversed(_visible_material_submissions(user, recycle_bin=recycle_bin)):
        creator_id = (submission.created_by or "").strip()
        if not creator_id or creator_id in seen:
            continue
        seen.add(creator_id)
        items.append({"id": creator_id, "name": _owner_name(creator_id) or creator_id})
    return {"items": items}


@router.get("/admin/material-submissions")
def list_material_submissions(page: int = Query(1, ge=1), size: int = Query(10, ge=1, le=100),
                              team_name: str = Query(""), drama_name: str = Query(""),
                              video_file_name: str = Query(""), title_name: str = Query(""),
                              can_upload_status: str = Query(""),
                              designated_upload_account_name: str = Query(""),
                              upload_account_name: str = Query(""),
                              created_by: str = Query(""), publish_status: str = Query(""),
                              sort_by: str = Query("created_time"), sort_order: str = Query("desc"),
                              user: dict = Depends(_user)):
    _require_auth(user)
    sort_by, sort_order = _validate_material_submission_sort(sort_by, sort_order)
    can_upload_value, can_upload_empty = _status_filter_arg(can_upload_status, kind="可上传状态")
    publish_value, publish_empty = _status_filter_arg(publish_status, kind="发布状态")
    visible_to_user_id = "" if user.get("role") == "admin" else user.get("id", "")
    filters = dict(
        team_name=team_name, drama_name=drama_name, video_file_name=video_file_name,
        title_name=title_name, can_upload_status=can_upload_value,
        can_upload_status_empty=can_upload_empty,
        designated_upload_account_name=designated_upload_account_name,
        upload_account_name=upload_account_name,
        created_by=created_by,
        publish_status=publish_value, publish_status_empty=publish_empty,
        visible_to_user_id=visible_to_user_id,
    )
    off, lim = _page_args(page, size)
    total = deps.material_submission_repo.count(**filters)
    page_items = deps.material_submission_repo.list(
        offset=off, limit=lim, sort_by=sort_by, sort_order=sort_order, **filters,
    )
    return _page_out([_submission_out(item, user) for item in page_items], total, page, size, key="submissions")


@router.get("/admin/material-submissions/recycle-bin")
def list_deleted_material_submissions(
    page: int = Query(1, ge=1), size: int = Query(10, ge=1, le=100),
    team_name: str = Query(""), drama_name: str = Query(""),
    video_file_name: str = Query(""), title_name: str = Query(""),
    can_upload_status: str = Query(""), designated_upload_account_name: str = Query(""),
    upload_account_name: str = Query(""), created_by: str = Query(""),
    publish_status: str = Query(""), sort_by: str = Query("created_time"),
    sort_order: str = Query("desc"), user: dict = Depends(_user),
):
    """列出仍在 7 天保留期内、且尚未完成 OSS 清理的逻辑删除提报。"""
    _require_auth(user)
    sort_by, sort_order = _validate_material_submission_sort(sort_by, sort_order)
    can_upload_value, can_upload_empty = _status_filter_arg(can_upload_status, kind="可上传状态")
    publish_value, publish_empty = _status_filter_arg(publish_status, kind="发布状态")
    visible_to_user_id = "" if user.get("role") == "admin" else user.get("id", "")
    filters = dict(
        team_name=team_name, drama_name=drama_name, video_file_name=video_file_name,
        title_name=title_name, can_upload_status=can_upload_value,
        can_upload_status_empty=can_upload_empty,
        designated_upload_account_name=designated_upload_account_name,
        upload_account_name=upload_account_name, created_by=created_by,
        publish_status=publish_value, publish_status_empty=publish_empty, recycle_bin=True,
        visible_to_user_id=visible_to_user_id,
    )
    off, lim = _page_args(page, size)
    total = deps.material_submission_repo.count(**filters)
    page_items = deps.material_submission_repo.list(
        offset=off, limit=lim, sort_by=sort_by, sort_order=sort_order, **filters,
    )
    return _page_out(
        [_submission_out(item, user) for item in page_items],
        total, page, size, key="submissions",
    )


@router.get("/admin/material-submissions/recycle-bin/{submission_id}")
def get_deleted_material_submission_detail(submission_id: str, user: dict = Depends(_user)):
    """读取尚未进入 OSS 清理流程的回收站提报。"""
    return _submission_out(_require_deleted_submission_access(user, submission_id), user)


@router.get("/admin/material-submissions/recycle-bin/{submission_id}/operations")
def get_deleted_material_submission_operations(submission_id: str, user: dict = Depends(_user)):
    submission = _require_deleted_submission_access(user, submission_id)
    return _material_submission_operations_out(submission)


@router.put("/admin/material-submissions/permissions/batch")
def batch_set_material_submission_permissions(
    body: schemas.MaterialSubmissionBatchPermissionsIn,
    user: dict = Depends(_user),
):
    _require_perm(user, "admin.grant")
    permission_type = (body.permission_type or "").strip()
    if permission_type not in ("read", "read_edit"):
        raise HTTPException(400, "权限类型只能是 read 或 read_edit")
    submission_ids = list(dict.fromkeys(str(item or "").strip() for item in body.submission_ids))
    submission_ids = [item for item in submission_ids if item]
    user_ids = list(dict.fromkeys(str(item or "").strip() for item in body.user_ids))
    user_ids = [item for item in user_ids if item]
    if not submission_ids:
        raise HTTPException(400, "请至少选择一个素材提报")
    if not user_ids:
        raise HTTPException(400, "请至少选择一个用户")
    if len(submission_ids) > 1000 or len(user_ids) > 200:
        raise HTTPException(400, "单次分配数量过多")
    known_users = {account.id: account for account in deps.user_repo.list()}
    targets = []
    for user_id in user_ids:
        account = known_users.get(user_id)
        if account is None:
            raise HTTPException(400, f"用户不存在: {user_id}")
        if account.role == "admin":
            raise HTTPException(400, "管理员默认拥有全部数据权限，无需分配")
        targets.append(account)
    for submission_id in submission_ids:
        submission = _require_submission_access(user, submission_id)
        before = deps.material_submission_repo.permissions_for(submission_id)
        grants = dict(before)
        for account in targets:
            grants[account.id] = permission_type
        owner = known_users.get(submission.created_by)
        if owner is not None and owner.role != "admin":
            grants[owner.id] = "read_edit"
        deps.material_submission_repo.replace_permissions(submission_id, grants, by=user["id"])
        changes = _submission_permission_changes(
            before, deps.material_submission_repo.permissions_for(submission_id),
        )
        if changes:
            _record_submission_operation(submission_id, "permission", user, changes)
    return {
        "submission_count": len(submission_ids),
        "user_count": len(targets),
        "grant_count": len(submission_ids) * len(targets),
        "permission_type": permission_type,
    }


def _material_submission_permission_target(user_id: str):
    account = deps.user_repo.get(user_id)
    if account is None:
        raise HTTPException(404, "用户不存在")
    if account.role == "admin":
        raise HTTPException(400, "管理员默认拥有全部数据权限，无需分配")
    return account


@router.get("/admin/material-submissions/permissions/user/{target_user_id}")
def get_material_submission_user_permissions(target_user_id: str, user: dict = Depends(_user)):
    _require_perm(user, "admin.grant")
    account = _material_submission_permission_target(target_user_id)
    grants = []
    permission_by_submission = deps.material_submission_repo.permissions_for_user(account.id)
    submissions = deps.material_submission_repo.list(
        offset=0, limit=None, sort_by="created_time", sort_order="desc",
    )
    for submission in submissions:
        is_owner = submission.created_by == account.id
        permission_type = "read_edit" if is_owner else permission_by_submission.get(submission.id, "")
        if permission_type in ("read", "read_edit"):
            grants.append({
                "submission_id": submission.id,
                "permission_type": permission_type,
                "locked": is_owner,
                "submission": _submission_out(submission, user),
            })
    return {
        "user": {"id": account.id, "name": account.name},
        "grants": grants,
    }


_REQUIREMENT_URGENCIES = {"low", "medium", "high"}
_REQUIREMENT_STATUSES = {
    "not_started", "pending_reply", "in_progress", "pending_acceptance", "completed",
    "acceptance_failed",
}


def _check_requirement_values(urgency: str, status: str) -> tuple[str, str]:
    urgency = (urgency or "").strip()
    status = (status or "").strip()
    if urgency not in _REQUIREMENT_URGENCIES:
        raise HTTPException(400, "非法紧急程度: 应为 low/medium/high")
    if status not in _REQUIREMENT_STATUSES:
        raise HTTPException(400, "非法需求状态")
    return urgency, status


def _requirement_out(item: Requirement, user: dict) -> dict:
    can_edit = user.get("role") == "admin" or item.created_by == user.get("id")
    return {
        "id": item.id, "description": item.description, "urgency": item.urgency,
        "status": item.status, "reply": item.reply, "attachments": list(item.attachments or []),
        "created_by": item.created_by, "created_by_name": _owner_name(item.created_by),
        "created_time": item.created_time, "updated_by": item.updated_by,
        "updated_by_name": _owner_name(item.updated_by), "updated_time": item.updated_time,
        "can_edit": can_edit,
    }


def _require_requirement(requirement_id: str, user: dict, edit: bool = False) -> Requirement:
    _require_auth(user)
    item = deps.requirement_repo.get(requirement_id)
    if item is None:
        raise HTTPException(404, "需求不存在")
    if edit and user.get("role") != "admin" and item.created_by != user.get("id"):
        raise HTTPException(403, "无权修改该需求")
    return item


@router.get("/requirements")
def list_requirements(q: str = "", urgency: str = "", status: str = "",
                      page: int = Query(1, ge=1), size: int = Query(20, ge=1, le=100),
                      user: dict = Depends(_user)):
    _require_auth(user)
    if urgency and urgency not in _REQUIREMENT_URGENCIES:
        raise HTTPException(400, "非法紧急程度筛选值")
    if status and status not in _REQUIREMENT_STATUSES:
        raise HTTPException(400, "非法需求状态筛选值")
    offset, limit = _page_args(page, size)
    items = deps.requirement_repo.list(q=q.strip(), urgency=urgency, status=status,
                                       offset=offset, limit=limit)
    total = deps.requirement_repo.count(q=q.strip(), urgency=urgency, status=status)
    return _page_out([_requirement_out(x, user) for x in items], total, page, size, key="requirements")


@router.post("/requirements")
def create_requirement(body: schemas.RequirementIn, user: dict = Depends(_user)):
    _require_auth(user)
    urgency, status = _check_requirement_values(body.urgency, body.status)
    attachments = [str(x).strip() for x in body.attachments if str(x).strip()]
    item = Requirement(id=next_id_str(), description=body.description.strip(), urgency=urgency,
                       status=status, reply=body.reply.strip(), attachments=attachments,
                       created_by=user["id"])
    deps.requirement_repo.add(item, by=user["id"])
    return _requirement_out(deps.requirement_repo.get(item.id) or item, user)


@router.get("/requirements/{requirement_id}")
def get_requirement(requirement_id: str, user: dict = Depends(_user)):
    return _requirement_out(_require_requirement(requirement_id, user), user)


@router.put("/requirements/{requirement_id}")
def update_requirement(requirement_id: str, body: schemas.RequirementIn,
                       user: dict = Depends(_user)):
    current = _require_requirement(requirement_id, user, edit=True)
    urgency, status = _check_requirement_values(body.urgency, body.status)
    current.description = body.description.strip()
    current.urgency = urgency
    current.status = status
    current.reply = body.reply.strip()
    current.attachments = [str(x).strip() for x in body.attachments if str(x).strip()]
    deps.requirement_repo.add(current, by=user["id"])
    return _requirement_out(deps.requirement_repo.get(current.id) or current, user)


@router.delete("/requirements/{requirement_id}")
def delete_requirement(requirement_id: str, user: dict = Depends(_user)):
    _require_requirement(requirement_id, user, edit=True)
    deps.requirement_repo.delete(requirement_id, by=user["id"])
    return {"deleted": requirement_id}


@router.post("/admin/material-submissions/permissions/user/{target_user_id}/unselected")
def list_unselected_material_submission_permissions(
    target_user_id: str,
    body: schemas.MaterialSubmissionUnselectedQueryIn,
    user: dict = Depends(_user),
):
    """按当前未保存的权限草稿分页，确保移入/移出后当前页自动补位。"""
    _require_perm(user, "admin.grant")
    _material_submission_permission_target(target_user_id)
    selected_ids = {
        str(item or "").strip() for item in body.selected_submission_ids if str(item or "").strip()
    }
    if len(selected_ids) > 1000:
        raise HTTPException(400, "单次分配数量过多")

    can_upload_value, can_upload_empty = _status_filter_arg(
        body.can_upload_status, kind="可上传状态"
    )
    publish_value, publish_empty = _status_filter_arg(body.publish_status, kind="发布状态")
    sort_by, sort_order = _validate_material_submission_sort(body.sort_by, body.sort_order)
    filters = dict(
        team_name=body.team_name, drama_name=body.drama_name,
        title_name=body.title_name,
        can_upload_status=can_upload_value, can_upload_status_empty=can_upload_empty,
        designated_upload_account_name=body.designated_upload_account_name,
        upload_account_name=body.upload_account_name, created_by=body.created_by,
        publish_status=publish_value,
        publish_status_empty=publish_empty,
        exclude_ids=selected_ids,
    )
    total = deps.material_submission_repo.count(**filters)
    off, lim = _page_args(body.page, body.size)
    page_items = deps.material_submission_repo.list(
        offset=off, limit=lim, sort_by=sort_by, sort_order=sort_order, **filters,
    )
    return _page_out(
        [_submission_out(item, user) for item in page_items],
        total, body.page, body.size, key="submissions",
    )


@router.put("/admin/material-submissions/permissions/user/{target_user_id}")
def set_material_submission_user_permissions(
    target_user_id: str,
    body: schemas.MaterialSubmissionUserPermissionsIn,
    user: dict = Depends(_user),
):
    _require_perm(user, "admin.grant")
    account = _material_submission_permission_target(target_user_id)
    desired: dict[str, str] = {}
    for grant in body.grants:
        submission_id = (grant.submission_id or "").strip()
        permission_type = (grant.permission_type or "").strip()
        if not submission_id:
            continue
        if permission_type not in ("read", "read_edit"):
            raise HTTPException(400, "权限类型只能是 read 或 read_edit")
        desired[submission_id] = permission_type
    if len(desired) > 1000:
        raise HTTPException(400, "单次分配数量过多")

    submissions = deps.material_submission_repo.list(offset=0, limit=None)
    known_submission_ids = {item.id for item in submissions}
    missing_ids = sorted(set(desired) - known_submission_ids)
    if missing_ids:
        raise HTTPException(400, f"素材提报不存在: {missing_ids[0]}")

    before = deps.material_submission_repo.permissions_for_user(account.id)
    changed_count = deps.material_submission_repo.replace_user_permissions(
        account.id, desired, by=user["id"]
    )
    after = deps.material_submission_repo.permissions_for_user(account.id)
    for submission in submissions:
        old_value = before.get(submission.id, "")
        new_value = after.get(submission.id, "")
        if old_value == new_value:
            continue
        _record_submission_operation(submission.id, "permission", user, [{
            "field": "permission_type",
            "user_id": account.id,
            "user_name": account.name,
            "before": old_value,
            "after": new_value,
        }])
    effective = dict(desired)
    for submission in submissions:
        if submission.created_by == account.id:
            effective[submission.id] = "read_edit"
    grants = [
        {
            "submission_id": submission.id,
            "permission_type": effective[submission.id],
            "locked": submission.created_by == account.id,
        }
        for submission in submissions if submission.id in effective
    ]
    return {
        "user": {"id": account.id, "name": account.name},
        "grants": grants,
        "changed_count": changed_count,
    }

@router.get("/admin/material-submissions/{submission_id}")
def get_material_submission_detail(submission_id: str, user: dict = Depends(_user)):
    return _submission_out(_require_submission_access(user, submission_id), user)


@router.post("/admin/material-submissions/{submission_id}/decode", status_code=202)
def start_material_submission_decode(submission_id: str, user: dict = Depends(_user)):
    submission = _require_submission_access(user, submission_id)
    if not submission.oss_key:
        raise HTTPException(400, "该提报没有源视频")
    if submission.decoded_oss_key:
        return _submission_decode_status_out(submission, user)
    if submission.requires_decode != 1:
        return {
            "submission_id": submission.id,
            "status": "not_required",
            "stage": "done",
            "source_key": submission.oss_key,
            "decoded_oss_key": "",
            "decoded_video_url": "",
            "error": "",
            "progress": 100,
        }

    coordinator = _submission_decode_coordinator()
    try:
        existing_status = coordinator.get_status(submission.id)
    except Exception as exc:
        raise HTTPException(503, "Redis 解码状态服务不可用") from exc
    if (
        existing_status
        and existing_status.get("status") == "not_required"
        and existing_status.get("source_key") == submission.oss_key
    ):
        return existing_status
    try:
        distributed_lock = coordinator.acquire(
            submission.id,
            timeout_seconds=settings.submission_decode_lock_ttl_seconds,
        )
    except Exception as exc:
        raise HTTPException(503, "Redis 解码锁服务不可用") from exc

    if distributed_lock is None:
        status = _submission_decode_status_out(submission, user)
        if status.get("status") == "idle":
            status.update(status="processing", stage="queued")
        return status

    try:
        _set_submission_decode_status(submission.id, "queued", "queued", error="", progress=1)
        deps.submission_transcode_pool.submit(
            _run_submission_decode_job,
            submission.id,
            submission.oss_key,
            submission.video_file_name,
            user.get("id", ""),
            distributed_lock,
        )
    except Exception as exc:
        try:
            distributed_lock.release()
        except Exception:
            pass
        raise HTTPException(503, "解码任务提交失败") from exc
    return _submission_decode_status_out(submission, user)


@router.get("/admin/material-submissions/{submission_id}/decode")
def get_material_submission_decode_status(submission_id: str, user: dict = Depends(_user)):
    submission = _require_submission_access(user, submission_id)
    return _submission_decode_status_out(submission, user)


def _material_submission_operations_out(submission: MaterialSubmission) -> dict:
    operations = deps.material_submission_repo.list_operations(submission.id)
    if not any(item.get("action") == "create" for item in operations):
        operations.append({
            "id": f"created-{submission.id}",
            "action": "create",
            "operator_id": submission.created_by,
            "changes": [],
            "operation_time": submission.created_time,
        })
    for item in operations:
        item["operator_name"] = _owner_name(item.get("operator_id", ""))
    operations.sort(key=lambda item: str(item.get("operation_time", "")), reverse=True)
    return {"submission_id": submission.id, "operations": operations}


@router.get("/admin/material-submissions/{submission_id}/operations")
def get_material_submission_operations(submission_id: str, user: dict = Depends(_user)):
    submission = _require_submission_access(user, submission_id)
    return _material_submission_operations_out(submission)


@router.get("/admin/material-submissions/{submission_id}/permissions")
def get_material_submission_permissions(submission_id: str, user: dict = Depends(_user)):
    _require_perm(user, "admin.grant")
    submission = _require_submission_access(user, submission_id)
    grants = deps.material_submission_repo.permissions_for(submission_id)
    users = []
    for account in sorted(deps.user_repo.list(), key=lambda item: item.name.lower()):
        is_admin = account.role == "admin"
        is_owner = account.id == submission.created_by
        permission_type = "read_edit" if (is_admin or is_owner) else grants.get(account.id, "")
        users.append({
            "user_id": account.id, "name": account.name, "role": account.role,
            "permission_type": permission_type, "is_owner": is_owner, "locked": is_admin or is_owner,
        })
    return {"submission_id": submission_id, "users": users}


@router.put("/admin/material-submissions/{submission_id}/permissions")
def set_material_submission_permissions(submission_id: str, body: schemas.MaterialSubmissionPermissionsIn,
                                        user: dict = Depends(_user)):
    _require_perm(user, "admin.grant")
    submission = _require_submission_access(user, submission_id)
    before = deps.material_submission_repo.permissions_for(submission_id)
    known_users = {account.id: account for account in deps.user_repo.list()}
    grants: dict[str, str] = {}
    for grant in body.grants:
        user_id = (grant.user_id or "").strip()
        permission_type = (grant.permission_type or "").strip()
        account = known_users.get(user_id)
        if account is None:
            raise HTTPException(400, f"用户不存在: {user_id}")
        if permission_type not in ("read", "read_edit"):
            raise HTTPException(400, "权限类型只能是 read 或 read_edit")
        if account.role != "admin":
            grants[user_id] = permission_type
    owner = known_users.get(submission.created_by)
    if owner is not None and owner.role != "admin":
        grants[submission.created_by] = "read_edit"
    deps.material_submission_repo.replace_permissions(submission_id, grants, by=user["id"])
    changes = _submission_permission_changes(
        before, deps.material_submission_repo.permissions_for(submission_id),
    )
    if changes:
        _record_submission_operation(submission_id, "permission", user, changes)
    return get_material_submission_permissions(submission_id, user)

def _submission_in_to_model(body: schemas.MaterialSubmissionIn, *, sid: str, by: str) -> MaterialSubmission:
    decoded_oss_key = (body.decoded_oss_key or "").strip()
    return MaterialSubmission(
        id=sid,
        team_name=(body.team_name or "").strip(),
        delivery_time=(body.delivery_time or "").strip(),
        drama_name=(body.drama_name or "").strip(),
        oss_key=(body.oss_key or "").strip(),
        decoded_oss_key=decoded_oss_key,
        requires_decode=1 if decoded_oss_key else 0,
        video_file_name=(body.video_file_name or "").strip(),
        title_name=(body.title_name or "").strip(),
        episode_range=(body.episode_range or "").strip(),
        upload_date=(body.upload_date or "").strip(),
        created_by=by,
    )


def _submission_update_to_model(body: schemas.MaterialSubmissionUpdateIn, *, sid: str, by: str) -> MaterialSubmission:
    can_upload_status = _check_can_upload_status(body.can_upload_status)
    publish_status = _check_publish_status(body.publish_status)
    attachments = []
    for x in body.platform_reject_attachments or []:
        v = (x or "").strip()
        if v:
            attachments.append(v)
    decoded_oss_key = (body.decoded_oss_key or "").strip()
    return MaterialSubmission(
        id=sid,
        team_name=(body.team_name or "").strip(),
        delivery_time=(body.delivery_time or "").strip(),
        drama_name=(body.drama_name or "").strip(),
        oss_key=(body.oss_key or "").strip(),
        decoded_oss_key=decoded_oss_key,
        requires_decode=1 if decoded_oss_key else 0,
        video_file_name=(body.video_file_name or "").strip(),
        title_name=(body.title_name or "").strip(),
        episode_range=(body.episode_range or "").strip(),
        revision_comment=(body.revision_comment or "").strip(),
        can_upload_status=can_upload_status,
        designated_upload_account_name=(body.designated_upload_account_name or "").strip(),
        upload_account_name=(body.upload_account_name or "").strip(),
        upload_date=(body.upload_date or "").strip(),
        publish_status=publish_status,
        platform_reject_reason=(body.platform_reject_reason or "").strip(),
        platform_reject_attachments=attachments,
        created_by=by,
    )


def _apply_submission_process_fields(
    submission: MaterialSubmission,
    body: schemas.MaterialSubmissionProcessIn,
) -> MaterialSubmission:
    can_upload_status = _check_can_upload_status(body.can_upload_status)
    publish_status = _check_publish_status(body.publish_status)
    attachments = []
    for x in body.platform_reject_attachments or []:
        v = (x or "").strip()
        if v:
            attachments.append(v)
    submission.revision_comment = (body.revision_comment or "").strip()
    submission.can_upload_status = can_upload_status
    submission.designated_upload_account_name = (body.designated_upload_account_name or "").strip()
    submission.upload_account_name = (body.upload_account_name or "").strip()
    if body.upload_date is not None:
        submission.upload_date = body.upload_date.strip()
    submission.publish_status = publish_status
    submission.platform_reject_reason = (body.platform_reject_reason or "").strip()
    submission.platform_reject_attachments = attachments
    return submission


@router.post("/admin/material-submissions")
def create_material_submission(body: schemas.MaterialSubmissionIn, user: dict = Depends(_user)):
    _require_auth(user)
    s = _submission_in_to_model(body, sid=next_id_str(), by=user["id"])
    deps.material_submission_repo.add(s, by=user["id"])
    _record_submission_operation(s.id, "create", user, [])
    if user.get("role") != "admin":
        deps.material_submission_repo.replace_permissions(s.id, {user["id"]: "read_edit"}, by=user["id"])
    return _submission_out(deps.material_submission_repo.get(s.id) or s, user)


@router.put("/admin/material-submissions/{submission_id}")
def update_material_submission(submission_id: str, body: schemas.MaterialSubmissionUpdateIn,
                               user: dict = Depends(_user)):
    cur = _require_submission_access(user, submission_id, edit=True)
    s = _submission_update_to_model(body, sid=submission_id, by=cur.created_by or user["id"])
    # 只有调用方完全未提交处理字段时才保留旧值；明确提交 null/空串表示清空。
    fields_set = getattr(body, "model_fields_set", None)
    if fields_set is None:  # Pydantic v1 compatibility
        fields_set = getattr(body, "__fields_set__", set())
    process_fields = {
        "revision_comment", "can_upload_status", "designated_upload_account_name",
        "upload_account_name", "upload_date",
        "publish_status", "platform_reject_reason", "platform_reject_attachments",
    }
    process_fields_omitted = not (set(fields_set) & process_fields)
    if "decoded_oss_key" not in fields_set:
        s.decoded_oss_key = cur.decoded_oss_key
        s.requires_decode = cur.requires_decode
    if process_fields_omitted:
        s.revision_comment = cur.revision_comment
        s.can_upload_status = cur.can_upload_status
        s.designated_upload_account_name = cur.designated_upload_account_name
        s.upload_account_name = cur.upload_account_name
        s.upload_date = cur.upload_date
        s.publish_status = cur.publish_status
        s.platform_reject_reason = cur.platform_reject_reason
        s.platform_reject_attachments = list(cur.platform_reject_attachments or [])
    elif "upload_date" not in fields_set:
        s.upload_date = cur.upload_date
    deps.material_submission_repo.add(s, by=user["id"])
    changes = _submission_changes(cur, s)
    if changes:
        _record_submission_operation(s.id, "update", user, changes)
    return _submission_out(deps.material_submission_repo.get(s.id) or s, user)


@router.put("/admin/material-submissions/{submission_id}/process")
def process_material_submission(submission_id: str, body: schemas.MaterialSubmissionProcessIn,
                                user: dict = Depends(_user)):
    cur = _require_submission_access(user, submission_id, edit=True)
    before = MaterialSubmission(**vars(cur))
    cur = _apply_submission_process_fields(cur, body)
    deps.material_submission_repo.add(cur, by=user["id"])
    changes = _submission_changes(before, cur)
    if changes:
        _record_submission_operation(cur.id, "process", user, changes)
    return _submission_out(deps.material_submission_repo.get(cur.id) or cur, user)


@router.post("/admin/material-submissions/batch/delete")
def delete_material_submissions(body: schemas.IdsIn, user: dict = Depends(_user)):
    _require_auth(user)
    deleted = []
    for rid in body.ids:
        _require_submission_access(user, rid, edit=True)
        if deps.material_submission_repo.delete(rid, by=user["id"]):
            _record_submission_operation(
                rid, "delete", user,
                [{"field": "deleted", "before": False, "after": True}],
            )
            deleted.append(rid)
    return {"deleted": deleted}


@router.post("/admin/material-submissions/{submission_id}/restore")
def restore_material_submission(submission_id: str, user: dict = Depends(_user)):
    _require_deleted_submission_access(user, submission_id, edit=True)
    if not deps.material_submission_repo.restore(submission_id, by=user["id"]):
        raise HTTPException(409, "该素材提报已恢复或已完成 OSS 清理")
    _record_submission_operation(
        submission_id, "restore", user,
        [{"field": "deleted", "before": True, "after": False}],
    )
    restored = deps.material_submission_repo.get(submission_id)
    if restored is None:
        raise HTTPException(409, "素材提报恢复失败")
    return _submission_out(restored, user)


# ── 作品审核记录(管理员):只作品,按项目分组,按提交时间(AuditTask.created_ms)区间筛 ──
def _work_out(task: AuditTask) -> dict:
    m = deps.material_repo.get(task.material_id) if task.material_id else None
    status = getattr(m.audit_status, "value", m.audit_status) if m else (task.verdict or "review")
    rejects = list(getattr(m, "reject_events", []) or []) if m else []
    link = ""
    if m and m.oss_key:
        try:
            link = deps.storage.download_url(m.oss_key)
        except Exception:
            link = ""
    return {"task_id": task.id, "name": task.name, "owner_name": _owner_name(task.owner_id),
            "created_ms": task.created_ms, "status": status,
            "reject_count": len(rejects), "reject_events": rejects,
            "report_id": task.report_id, "project_id": getattr(task, "project_id", ""),
            "download_url": link}


def _collect_works(from_ms, to_ms, status):
    """所有作品(video_kind=work)的审核记录,按提交时间区间 + 最终状态筛,按时间倒序。"""
    out = []
    for t in deps.task_repo.list_all():
        if getattr(t, "video_kind", "material") != "work":
            continue
        if from_ms is not None and t.created_ms < from_ms:
            continue
        if to_ms is not None and t.created_ms >= to_ms:      # 半开区间 [from, to)
            continue
        w = _work_out(t)
        if status is not None and w["status"] != status:
            continue
        out.append(w)
    out.sort(key=lambda w: w["created_ms"], reverse=True)
    return out


def _fmt_local(ms, tz_offset_min, date_only=False):
    """epoch 毫秒 → 浏览器本地时间字符串(前端传 getTimezoneOffset(),国内=-480)。"""
    if not ms:
        return ""
    local = (int(ms) - int(tz_offset_min) * 60000) / 1000.0
    return time.strftime("%Y%m%d" if date_only else "%Y-%m-%d %H:%M", time.gmtime(local))


@router.get("/works")
def list_works(from_ms: int | None = None, to_ms: int | None = None,
               status: str | None = None, user: dict = Depends(_user)):
    """作品审核记录,按项目分组。可选提交时间区间 [from_ms, to_ms) + 最终状态。仅管理员。"""
    _require_perm(user, "materials.audit")
    works = _collect_works(from_ms, to_ms, _check_status(status))
    groups: dict[str, dict] = {}
    for w in works:
        pid = w["project_id"]
        g = groups.get(pid)
        if g is None:
            proj = deps.project_repo.get(pid)
            g = groups[pid] = {"project_id": pid, "project_name": (proj.name if proj else "(未归属)"),
                               "count": 0, "works": []}
        g["works"].append(w)
        g["count"] += 1
    return {"groups": list(groups.values()), "total": len(works)}


@router.get("/works/export.xlsx")
def export_works(from_ms: int | None = None, to_ms: int | None = None, status: str | None = None,
                 tz_offset_min: int = 0, user: dict = Depends(_user)):
    """导出作品审核记录为 .xlsx。列一行内容,退回原因合并到一格。仅管理员。"""
    _require_perm(user, "materials.audit")
    import io
    from urllib.parse import quote
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    works = _collect_works(from_ms, to_ms, _check_status(status))
    st_cn = {"pass": "通过", "block": "退回", "review": "待复核"}
    wb = Workbook()
    ws = wb.active
    ws.title = "作品审核记录"
    ws.append(["项目", "作品名称", "上传者", "提交时间", "最终状态", "退回次数", "退回原因", "素材链接"])
    for w in works:
        proj = deps.project_repo.get(w["project_id"])
        reasons = "; ".join(f"{i + 1}) {(e.get('reason') or '')[:60]}"
                            for i, e in enumerate(w["reject_events"]))
        ws.append([proj.name if proj else "", w["name"], w["owner_name"],
                   _fmt_local(w["created_ms"], tz_offset_min), st_cn.get(w["status"], w["status"]),
                   w["reject_count"], reasons, w["download_url"]])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"作品审核记录-{_fmt_local(to_ms or int(time.time() * 1000), tz_offset_min, date_only=True)}.xlsx"
    return StreamingResponse(
        bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"})


# ── 规则训练(管理员:按项目标注样本 → AI 迭代调优项目规则)──
def _ts_out(ts) -> dict:
    return {
        "id": ts.id, "project_id": ts.project_id, "name": ts.name,
        "status": ts.status, "max_fp_ratio": ts.max_fp_ratio,
        "max_iterations": ts.max_iterations,
        "rule_snapshot": ts.rule_snapshot,
        "training_result": ts.training_result,
        "started_at": getattr(ts, "started_at", ""),
        "completed_at": getattr(ts, "completed_at", ""),
    }


def _te_out(te) -> dict:
    task_id = ""
    task_name = ""
    for t in deps.task_repo.list_all():
        if getattr(t, "material_id", "") == te.material_id:
            task_id = t.id
            task_name = getattr(t, "name", "") or ""
            break
    return {
        "id": te.id, "training_set_id": te.training_set_id,
        "material_id": te.material_id,
        "expected_rule_ids": te.expected_rule_ids,
        "source_note": te.source_note,
        "source_task_id": task_id,
        "source_task_name": task_name,
    }


@router.post("/training/projects/{project_id}/examples")
def add_training_example(project_id: str, body: schemas.TrainingExampleIn,
                         user: dict = Depends(_user)):
    """往项目训练集添加一条样本:标注某物料应被哪些规则命中。自动创建训练集(若不存在)。"""
    _require_perm(user, "audit.rules")
    try:
        te = deps.get_training_service().add_example(
            project_id=project_id, material_id=body.material_id,
            expected_rule_ids=body.expected_rule_ids,
            source_note=body.source_note, by=user["id"])
    except Exception as e:
        raise HTTPException(400, str(e))
    return _te_out(te)


@router.get("/training/projects/{project_id}/examples")
def list_training_examples(project_id: str, user: dict = Depends(_user)):
    """列某项目训练集的所有样本。"""
    _require_perm(user, "audit.rules")
    examples = deps.get_training_service().list_examples(project_id)
    return {"examples": [_te_out(e) for e in examples]}


@router.delete("/training/projects/{project_id}/examples/{example_id}")
def delete_training_example(project_id: str, example_id: str,
                            user: dict = Depends(_user)):
    """删除训练集里的一条样本。"""
    _require_perm(user, "audit.rules")
    deps.get_training_service().remove_example(example_id, by=user["id"])
    return {"deleted": example_id}


@router.put("/training/projects/{project_id}/examples/{example_id}")
def update_training_example(project_id: str, example_id: str,
                            body: schemas.TrainingExampleUpdateIn,
                            user: dict = Depends(_user)):
    """编辑训练样本:修改预期命中规则或标注备注。"""
    _require_perm(user, "audit.rules")
    try:
        te = deps.get_training_service().update_example(
            project_id, example_id,
            expected_rule_ids=body.expected_rule_ids,
            source_note=body.source_note, by=user["id"])
    except Exception as e:
        raise HTTPException(400, str(e))
    return _te_out(te)


@router.post("/training/projects/{project_id}/train")
def start_training(project_id: str, body: schemas.TrainingConfigIn | None = None,
                   user: dict = Depends(_user)):
    """启动规则训练(异步):快照当前规则 → 逐轮 AI 调优 → 重审 → 直到漏判=0 且多判率≤阈值。"""
    _require_perm(user, "audit.rules")
    fp = body.max_fp_ratio if body else None
    mi = body.max_iterations if body else None
    try:
        deps.get_training_service().start_training(
            project_id, by=user["id"], max_fp_ratio=fp, max_iterations=mi)
    except Exception as e:
        raise HTTPException(400, str(e))
    # 提交到审核线程池后台跑训练
    deps.audit_pool.submit(_run_training, project_id, user["id"])
    return {"status": "training", "project_id": project_id}


def _run_training(project_id: str, by: str) -> None:
    """后台执行训练循环。"""
    import logging
    _log = logging.getLogger(__name__)
    try:
        deps.get_training_service().run_training(project_id, by=by)
    except Exception as e:
        _log.exception("训练异常终止: project=%s, error=%s", project_id, e)


@router.get("/training/projects/{project_id}/status")
def get_training_status(project_id: str, user: dict = Depends(_user)):
    """获取项目训练集状态与结果。"""
    _require_perm(user, "audit.rules")
    ts = deps.get_training_service().get_status(project_id)
    if ts is None:
        return {"status": "not_created"}
    return _ts_out(ts)


@router.get("/training/sets")
def list_training_sets(page: int = Query(1, ge=1), size: int = Query(24, ge=1, le=100),
                      user: dict = Depends(_user)):
    """列出所有训练集(按创建时间倒序,分页)。"""
    _require_perm(user, "audit.rules")
    all_sets = deps.training_set_repo.list()
    total = len(all_sets)
    off, lim = _page_args(page, size)
    page_sets = all_sets[off:off + lim]
    return _page_out([_ts_out(ts) for ts in page_sets], total, page, size, key="sets")


@router.get("/training/sets/{ts_id}")
def get_training_set_detail(ts_id: str, user: dict = Depends(_user)):
    """获取训练集详情,含全部样本列表。"""
    _require_perm(user, "audit.rules")
    ts = deps.training_set_repo.get(ts_id)
    if ts is None:
        raise HTTPException(404, "训练集不存在")
    examples = deps.training_example_repo.list_for_set(ts.id)
    return {**_ts_out(ts), "examples": [_te_out(e) for e in examples]}


@router.delete("/training/sets/{ts_id}")
def delete_training_set(ts_id: str, user: dict = Depends(_user)):
    """删除训练集及其全部样本(软删)。"""
    _require_perm(user, "audit.rules")
    ts = deps.training_set_repo.get(ts_id)
    if ts is None:
        return {"deleted": ts_id}
    # 级联软删全部样本
    for te in deps.training_example_repo.list_for_set(ts.id):
        deps.training_example_repo.delete(te.id, by=user["id"])
    deps.training_set_repo.delete(ts_id, by=user["id"])
    return {"deleted": ts_id}
