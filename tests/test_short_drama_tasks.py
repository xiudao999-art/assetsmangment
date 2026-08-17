import io

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.api import deps
from app.domain.models import MaterialSubmission
from app.main import app

client = TestClient(app)


def _headers():
    token = client.post(
        "/users/login", json={"name": "demo", "password": "pw123456"}
    ).json()["token"]
    return {"Authorization": f"Bearer {token}"}


def _payload(name: str, **overrides):
    data = {
        "online_time": "2026-08-20",
        "drama_name": name,
        "task_type": "短剧",
        "tags": ["爆剧", "新剧"],
        "theme": "都市",
        "task_status": "未上线",
        "task_id": "TASK-001",
        "requirements": "按要求剪辑",
        "cover_oss_key": "short-drama-tasks/cover.png",
        "cloud_material_url": "https://pan.example.com/task",
        "topic_editing_requirements": "#话题",
        "submission_activity_time": "8月20日至9月20日",
        "settlement_mode": "按消耗分佣",
        "commission_validity_period": "30天",
        "settlement_period": "月结",
        "data_image_oss_key": "short-drama-tasks/data.png",
        "quality_case": "https://example.com/case",
        "remarks": "备注",
    }
    data.update(overrides)
    return data


def test_short_drama_task_image_upload_validates_and_returns_oss_key():
    headers = _headers()
    uploaded = client.post(
        "/admin/uploads/file",
        headers=headers,
        data={"scope": "short-drama-tasks"},
        files={"file": ("cover.png", b"test-image", "image/png")},
    )
    assert uploaded.status_code == 200, uploaded.text
    assert uploaded.json()["oss_key"].startswith("short-drama-tasks/")

    wrong_type = client.post(
        "/admin/uploads/file",
        headers=headers,
        data={"scope": "short-drama-tasks"},
        files={"file": ("notes.txt", b"not-an-image", "text/plain")},
    )
    assert wrong_type.status_code == 400

    oversized = client.post(
        "/admin/uploads/file",
        headers=headers,
        data={"scope": "short-drama-tasks"},
        files={"file": ("large.png", b"0" * (10 * 1024 * 1024 + 1), "image/png")},
    )
    assert oversized.status_code == 400


def test_short_drama_task_crud_and_unique_drama_name():
    headers = _headers()
    created = client.post(
        "/admin/short-drama-tasks", headers=headers, json=_payload("唯一剧名")
    )
    assert created.status_code == 200
    item = created.json()
    assert item["tags"] == ["爆剧", "新剧"]
    deps.material_submission_repo.add(MaterialSubmission(
        id="9001", team_name="团队甲", drama_name="唯一剧名", can_upload_status=1,
        publish_status=1,
    ))
    deps.material_submission_repo.add(MaterialSubmission(
        id="9002", team_name="团队乙", drama_name="唯一剧名-混剪", can_upload_status=1,
        publish_status=2,
    ))
    deps.material_submission_repo.add(MaterialSubmission(
        id="9003", team_name="团队甲", drama_name="唯一剧名", can_upload_status=2,
        publish_status=1,
    ))
    deps.material_submission_repo.add(MaterialSubmission(
        id="9004", team_name="其他团队", drama_name="无关剧名", can_upload_status=1,
        publish_status=1,
    ))
    detail_with_uploads = client.get(
        f"/admin/short-drama-tasks/{item['id']}", headers=headers,
    ).json()
    assert detail_with_uploads["upload_summary"] == {
        "team_names": ["团队乙", "团队甲"],
        "upload_count": 3,
        "can_upload_count": 2,
        "publish_success_count": 2,
    }
    assigned = client.put(
        f"/admin/short-drama-tasks/{item['id']}/pre-upload-teams", headers=headers,
        json={"team_names": ["团队甲", "团队乙", "团队甲"]},
    )
    assert assigned.status_code == 200
    assert assigned.json()["pre_upload_teams"] == ["团队甲", "团队乙"]
    type_options = client.get(
        "/admin/short-drama-tasks/options/task_type", headers=headers,
    )
    assert type_options.status_code == 200
    assert "短剧" in type_options.json()["items"]
    tag_options = client.get(
        "/admin/short-drama-tasks/options/tags", headers=headers,
        params={"keyword": "爆"},
    )
    assert tag_options.json()["items"] == ["爆剧"]
    team_options = client.get(
        "/admin/short-drama-tasks/options/pre_upload_teams", headers=headers,
        params={"keyword": "甲"},
    )
    assert team_options.json()["items"] == ["团队甲"]
    assert client.get(
        "/admin/short-drama-tasks/options/requirements", headers=headers,
    ).status_code == 400
    image_url = client.get(
        "/admin/uploads/url", headers=headers,
        params={"key": item["cover_oss_key"], "short_drama_task_id": item["id"]},
    )
    assert image_url.status_code == 200
    assert "url" in image_url.json()

    duplicate = client.post(
        "/admin/short-drama-tasks", headers=headers, json=_payload("唯一剧名")
    )
    assert duplicate.status_code == 409

    listed = client.get(
        "/admin/short-drama-tasks", headers=headers,
        params={
            "drama_name": "唯一", "task_status": "未上线", "task_id": "001",
            "online_time": "08-20", "tag": "爆", "pre_upload_team": "甲",
            "actual_upload_team": "乙",
        },
    ).json()
    assert listed["total"] == 1
    assert listed["tasks"][0]["upload_summary"]["upload_count"] == 3
    no_actual_team = client.get(
        "/admin/short-drama-tasks", headers=headers,
        params={"actual_upload_team": "不存在的团队"},
    ).json()
    assert no_actual_team["total"] == 0

    updated = client.put(
        f"/admin/short-drama-tasks/{item['id']}", headers=headers,
        json=_payload("唯一剧名", task_status="已上线", tags=["热播"]),
    )
    assert updated.status_code == 200
    assert updated.json()["task_status"] == "已上线"
    assert updated.json()["pre_upload_teams"] == ["团队甲", "团队乙"]

    assert client.delete(
        f"/admin/short-drama-tasks/{item['id']}", headers=headers
    ).status_code == 200
    assert client.get(
        f"/admin/short-drama-tasks/{item['id']}", headers=headers
    ).status_code == 404


def test_short_drama_task_excel_import_upserts_and_template():
    headers = _headers()
    existing = client.post(
        "/admin/short-drama-tasks", headers=headers, json=_payload("导入更新剧", remarks="旧值")
    )
    assert existing.status_code == 200

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["剧名", "任务状态", "标签", "备注", "网盘素材"])
    sheet.append(["导入更新剧", "已结束", '["完结","爆剧"]', "新值", "https://pan.example.com/a"])
    sheet.append(["导入新增剧", "未上线", "新剧、待播", "新增", "https://pan.example.com/b"])
    stream = io.BytesIO()
    workbook.save(stream)

    imported = client.post(
        "/admin/short-drama-tasks/import", headers=headers,
        files={"file": ("tasks.xlsx", stream.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert imported.status_code == 200, imported.text
    assert imported.json() == {"total": 2, "created": 1, "updated": 1}

    listed = client.get(
        "/admin/short-drama-tasks", headers=headers, params={"size": 100}
    ).json()
    by_name = {x["drama_name"]: x for x in listed["tasks"]}
    assert by_name["导入更新剧"]["remarks"] == "新值"
    assert by_name["导入更新剧"]["tags"] == ["完结", "爆剧"]
    client.put(
        f"/admin/short-drama-tasks/{by_name['导入更新剧']['id']}", headers=headers,
        json=_payload("导入更新剧", remarks="新值", tags=["完结", "爆剧"]),
    )

    sorted_list = client.get(
        "/admin/short-drama-tasks", headers=headers,
        params={"drama_name": "导入", "sort_by": "drama_name", "sort_order": "desc"},
    ).json()
    assert [x["drama_name"] for x in sorted_list["tasks"]] == ["导入更新剧", "导入新增剧"]

    template = client.get("/admin/short-drama-tasks/import-template", headers=headers)
    assert template.status_code == 200
    template_book = load_workbook(io.BytesIO(template.content), read_only=True)
    template_headers = next(template_book.active.values)
    assert "剧名" in template_headers
    assert "封面" in template_headers
    assert "数据图" in template_headers
