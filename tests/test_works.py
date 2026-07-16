"""作品审核记录:退回历史 + 按项目分组 + 按天筛选 + 导出 Excel(闭环③/④)。"""
import io
import time
from fastapi.testclient import TestClient
from app.main import app

client = TestClient(app)


def _tok(name, pw):
    return client.post("/users/login", json={"name": name, "password": pw}).json()["token"]


def _admin():
    return {"Authorization": f"Bearer {_tok('admin', 'admin123')}"}


def _user():
    return {"Authorization": f"Bearer {_tok('demo', 'pw123456')}"}


def _wait(task_id, hdr, n=150):
    for _ in range(n):
        t = client.get(f"/audit/tasks/{task_id}", headers=hdr).json()
        if t.get("status") in ("done", "failed"):
            return t
        time.sleep(0.02)
    return client.get(f"/audit/tasks/{task_id}", headers=hdr).json()


def _new_project(hdr, name):
    return client.post("/admin/projects", json={"name": name}, headers=hdr).json()["id"]


def _submit_work(hdr, pid, fname, data):
    r = client.post("/audit/submit", data={"type": "video", "video_kind": "work", "project_id": pid},
                    files={"file": (fname, data, "video/mp4")}, headers=hdr)
    return r.json()["task_id"]


# ── 退回历史记录(只人工拒绝算退回;机器只转人工、不记退回)──
def test_set_audit_block_records_reject_reason():
    from app.api import deps
    ah, uh = _admin(), _user()
    mid = client.post("/materials", json={"type": "image", "oss_key": "rej1.png"}, headers=uh).json()["id"]
    r = client.post(f"/materials/{mid}/set-audit", json={"status": "block", "reason": "画面血腥"}, headers=ah)
    assert r.status_code == 200
    ev = deps.material_repo.get(mid).reject_events
    assert len(ev) == 1 and ev[0]["reason"] == "画面血腥" and ev[0]["by"] == "人工" and ev[0]["ms"] > 0


def test_machine_flags_review_only_human_reject_records():
    """机器命中(含绝对禁词)只转「待审核」、不记退回;人工拒绝才拦截 + 记退回(by=人工)。"""
    from app.api import deps
    ah, uh = _admin(), _user()
    client.post("/admin/blockwords", json={"words": ["赌博"]}, headers=ah)
    r = client.post("/audit/submit", data={"type": "corpus", "content": "这是赌博广告机审要退回"}, headers=uh)
    t = _wait(r.json()["task_id"], uh)
    assert t["verdict"] == "review"                              # 机审只转人工,永不直接拦
    mid = t["material_id"]
    assert deps.material_repo.get(mid).reject_events == []       # 机审不记退回历史
    client.post(f"/materials/{mid}/set-audit", json={"status": "block", "reason": "人工确认违规"}, headers=ah)
    ev = deps.material_repo.get(mid).reject_events
    assert len(ev) == 1 and ev[0]["by"] == "人工"                # 人工拒绝才记退回


# ── /works:按项目分组 + 日期区间筛选 ──
def test_works_grouped_by_project_and_date_filter():
    from app.api import deps
    ah, uh = _admin(), _user()
    pA = _new_project(ah, "项目甲QC")
    pB = _new_project(ah, "项目乙QC")
    tA = _submit_work(uh, pA, "a1.mp4", b"work-a1"); _wait(tA, uh)
    tB = _submit_work(uh, pB, "b1.mp4", b"work-b1"); _wait(tB, uh)
    now = int(time.time() * 1000); day = 86400000
    for tid, ms in ((tA, now - day), (tB, now)):        # 甲=昨天,乙=今天
        task = deps.task_repo.get(tid); task.created_ms = ms; deps.task_repo.save(task)
    r = client.get("/works", params={"from_ms": now - 1000, "to_ms": now + day}, headers=ah).json()
    groups = {g["project_id"]: g for g in r["groups"]}
    assert pB in groups and groups[pB]["count"] == 1            # 今天:只项目乙 1 个
    assert pA not in groups                                     # 昨天的甲不在区间
    w = groups[pB]["works"][0]
    assert w["name"] == "b1.mp4" and "reject_count" in w and "created_ms" in w


def test_works_requires_admin():
    assert client.get("/works", headers=_user()).status_code == 403


def test_works_export_xlsx():
    import openpyxl
    ah, uh = _admin(), _user()
    pid = _new_project(ah, "导出项目QC")
    _wait(_submit_work(uh, pid, "exp1.mp4", b"work-exp1"), uh)
    r = client.get("/works/export.xlsx", headers=ah)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert "attachment" in r.headers.get("content-disposition", "").lower()
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert "作品名称" in header and "上传者" in header and "退回次数" in header
    rows = [[str(c.value) for c in row] for row in ws.iter_rows(min_row=2)]
    assert any(any("exp1.mp4" in cell for cell in row) for row in rows)
